"""Focused P&L report — dedicated Excel workbook for ALCO/Treasury daily review.

Distinct from excel_export.py (full dashboard dump): this module produces a
narratively coherent P&L workbook with Cover, KPI Dashboard, Day-over-Day
Bridge, Attribution Waterfall, Monthly Series, Top Contributors, 3-Way
Margin (FTP), Fixed vs Floating, Budget vs Actual, Forecast Tracking,
Alerts, CoC Decomposition, Deal P&L, and a Diagnostics sheet. Each sheet is
isolated by a try/except so a single failure cannot abort the export.
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from cockpit.export.styles import (
    FILL_NEG,
    FILL_POS,
    FMT_CURRENCY,
    FMT_DATE,
    FMT_INT,
    FMT_PERCENT,
    NAVY,
    SEVERITY_BADGE_FILLS,
    SEVERITY_ROW_FILLS,
    add_autofilter,
    add_color_scale,
    add_color_scale_range,
    apply_format,
    apply_format_by_header,
    apply_sign_fill,
    autofit_columns,
    footer,
    freeze,
    hide_gridlines,
    section_header,
    set_column_widths,
    setup_print_layout,
    style_header_row,
    title_block,
    write_empty_state,
    write_header_row,
    write_jump_link,
)
from openpyxl.utils import get_column_letter


def export_pnl_report(
    dashboard_data: dict,
    output_path: Path | str,
    date_run: str = "",
    deals: Optional[pd.DataFrame] = None,
) -> Path | None:
    """Write a focused P&L workbook. Returns path on success, None on failure.

    Each sheet is wrapped in its own try/except; a failure on one sheet logs a
    warning and continues with the next.
    """
    output_path = Path(output_path)
    wb = Workbook()
    wb.remove(wb.active)  # discard the default blank sheet

    diagnostics: list[tuple[str, str]] = []

    for name, builder in (
        ("Cover", lambda: _write_cover(wb, dashboard_data, date_run)),
        ("KPI Dashboard", lambda: _write_kpi_dashboard(wb, dashboard_data, date_run)),
        ("Day-over-Day Bridge", lambda: _write_dod_bridge(wb, dashboard_data, date_run)),
        ("Attribution Waterfall", lambda: _write_attribution(wb, dashboard_data, date_run)),
        ("P&L Explain (MoM)", lambda: _write_attribution_mom(wb, dashboard_data, date_run)),
        ("Monthly Series", lambda: _write_monthly_series(wb, dashboard_data, date_run)),
        ("Daily Projection (MTD)", lambda: _write_daily_projection(wb, dashboard_data, date_run)),
        ("Realized Daily P&L", lambda: _write_realized_daily_pnl(wb, dashboard_data, date_run)),
        ("BOOK2 \u0394MTM - Day-over-Day", lambda: _write_book2_delta_dod(wb, dashboard_data, date_run)),
        ("BOOK2 \u0394MTM - MTD", lambda: _write_book2_delta_mtd(wb, dashboard_data, date_run)),
        ("BOOK1 Realized - MTD", lambda: _write_book1_realized_mtd(wb, dashboard_data, date_run)),
        ("Top Contributors", lambda: _write_top_contributors(wb, dashboard_data, date_run)),
        ("Strategy P&L", lambda: _write_strategy(wb, dashboard_data, date_run)),
        ("Hedge Effectiveness", lambda: _write_hedge_effectiveness(wb, dashboard_data, date_run)),
        ("3-Way Margin (FTP)", lambda: _write_ftp(wb, dashboard_data, date_run)),
        ("Fixed vs Floating", lambda: _write_fixed_float(wb, dashboard_data, date_run)),
        ("Budget vs Actual", lambda: _write_budget(wb, dashboard_data, date_run)),
        ("Forecast Tracking", lambda: _write_forecast_tracking(wb, dashboard_data, date_run)),
        ("Alerts", lambda: _write_alerts(wb, dashboard_data, date_run)),
        ("CoC Decomposition", lambda: _write_coc(wb, dashboard_data, date_run)),
        ("Deal P&L", lambda: _write_deal_pnl(wb, dashboard_data, deals, date_run)),
    ):
        try:
            builder()
            diagnostics.append((name, "OK"))
        except Exception as e:  # noqa: BLE001 - per-sheet isolation is the design
            diagnostics.append((name, f"SKIPPED: {e}"))
            print(f"[pnl-report] Sheet '{name}' skipped: {e}")

    _write_diagnostics(wb, diagnostics, date_run, data=dashboard_data, deals=deals)

    # Second pass: fill jump-links on Cover now that all sheets exist.
    _fill_cover_jump_links(wb)

    # Polish pass: consistent print layout + gridlines on report-style sheets.
    report_style_sheets = {"Cover", "KPI Dashboard", "Diagnostics"}
    for sheet in wb.worksheets:
        setup_print_layout(sheet, date_run=date_run)
        if sheet.title in report_style_sheets:
            hide_gridlines(sheet)

    try:
        wb.save(output_path)
        return output_path
    except Exception as e:  # noqa: BLE001
        print(f"[pnl-report] Save failed: {e}")
        return None


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_cover(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Cover")
    summary = data.get("summary", {}) or {}
    kpis = summary.get("kpis", {}) or {}

    row = title_block(ws, "P&L Daily Report", f"Run date: {date_run}", span_cols=4)

    row = section_header(ws, row, "Headline NII (12M forward)")
    row = write_header_row(ws, row, ["Shock", "Total", "Realized", "Forecast", "Realized %"])

    labels = [("shock_0", "Base (0 bp)"), ("shock_50", "+50 bp"), ("shock_wirp", "WIRP (Market)")]
    for key, label in labels:
        entry = kpis.get(key)
        if not entry:
            continue
        ws.cell(row=row, column=1, value=label)
        total = entry.get("total", 0) or 0
        total_cell = ws.cell(row=row, column=2, value=total)
        total_cell.number_format = FMT_CURRENCY
        apply_sign_fill(total_cell, total)
        ws.cell(row=row, column=3, value=entry.get("realized", 0) or 0).number_format = FMT_CURRENCY
        ws.cell(row=row, column=4, value=entry.get("forecast", 0) or 0).number_format = FMT_CURRENCY
        pct = entry.get("realized_pct")
        if pct is not None:
            ws.cell(row=row, column=5, value=float(pct) / 100.0).number_format = FMT_PERCENT
        row += 1

    delta = kpis.get("delta_50_0")
    if delta is not None:
        ws.cell(row=row, column=1, value="Δ (+50bp vs Base)").font = Font(bold=True)
        delta_cell = ws.cell(row=row, column=2, value=float(delta))
        delta_cell.number_format = FMT_CURRENCY
        apply_sign_fill(delta_cell, delta)
        row += 1
    row += 1

    locked = summary.get("locked_in_nii")
    if locked and locked.get("has_data"):
        row = section_header(ws, row, "Locked-in NII (Fixed-rate contribution)")
        ws.cell(row=row, column=1, value="Total NII")
        ws.cell(row=row, column=2, value=float(locked.get("total_nii", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        ws.cell(row=row, column=1, value="Locked (Fixed-rate)")
        ws.cell(row=row, column=2, value=float(locked.get("locked_nii", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        ws.cell(row=row, column=1, value="Locked % of Total")
        pct_val = float(locked.get("locked_pct", 0) or 0) / 100.0
        ws.cell(row=row, column=2, value=pct_val).number_format = FMT_PERCENT
        row += 1
        if not locked.get("pct_meaningful", True):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1, value="⚠ Total NII ≤ 0 — certainty-floor narrative N/A").font = Font(italic=True, color="808080")
            row += 1
        row += 1

    coc_ytd = summary.get("coc_ytd")
    if coc_ytd:
        row = section_header(ws, row, "Year-to-Date P&L")
        for label, key in (
            ("Gross Carry", "gross_carry"),
            ("Funding Cost", "funding_cost"),
            ("P&L (Simple)", "pnl_simple"),
            ("P&L (Compounded)", "pnl_compounded"),
        ):
            val = coc_ytd.get(key, 0) or 0
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(val)).number_format = FMT_CURRENCY
            row += 1
        row += 1

    dq = data.get("data_quality") or {}
    if dq:
        n_issues = len(dq.get("issues", [])) if isinstance(dq.get("issues"), list) else 0
        status = "Clean" if n_issues == 0 else f"{n_issues} issue(s)"
        ws.cell(row=row, column=1, value=f"Data Quality: {status}").font = Font(italic=True, color="595959")
        row += 2

    # Jump links filled in a second pass once all sheets exist; the anchor is
    # stashed on the worksheet so _fill_cover_jump_links can find it without
    # needing a separate registry. (openpyxl tolerates the extra attribute.)
    row = section_header(ws, row, "Jump to sheet")
    ws._cover_jump_links_anchor = row  # type: ignore[attr-defined]

    set_column_widths(ws, [34, 20, 20, 20, 16])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="summary")


def _write_kpi_dashboard(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("KPI Dashboard")
    summary = data.get("summary", {}) or {}
    kpis = summary.get("kpis", {}) or {}

    row = title_block(ws, "KPI Dashboard", "Per-shock NII + YTD reconciliation", span_cols=6)

    row = write_header_row(ws, row, ["shock", "total", "realized", "forecast", "unit", "realized_pct"])
    data_start = row
    for key in ("shock_0", "shock_50", "shock_wirp"):
        entry = kpis.get(key)
        if not entry:
            continue
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=float(entry.get("total", 0) or 0)).number_format = FMT_CURRENCY
        ws.cell(row=row, column=3, value=float(entry.get("realized", 0) or 0)).number_format = FMT_CURRENCY
        ws.cell(row=row, column=4, value=float(entry.get("forecast", 0) or 0)).number_format = FMT_CURRENCY
        ws.cell(row=row, column=5, value=entry.get("unit", "CHF"))
        pct = entry.get("realized_pct")
        if pct is not None:
            ws.cell(row=row, column=6, value=float(pct) / 100.0).number_format = FMT_PERCENT
        row += 1
    delta = kpis.get("delta_50_0")
    if delta is not None:
        ws.cell(row=row, column=1, value="delta_50_0").font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(delta)).number_format = FMT_CURRENCY
        row += 1

    row += 2
    coc_ytd = summary.get("coc_ytd")
    if coc_ytd:
        row = section_header(ws, row, "YTD Reconciliation (Simple vs Compounded)")
        row = write_header_row(ws, row, ["metric", "value"])
        for label, key in (
            ("Gross Carry", "gross_carry"),
            ("Funding Cost", "funding_cost"),
            ("P&L (Simple)", "pnl_simple"),
            ("P&L (Compounded)", "pnl_compounded"),
        ):
            val = coc_ytd.get(key, 0) or 0
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=float(val)).number_format = FMT_CURRENCY
            row += 1

    row += 2
    locked = summary.get("locked_in_nii")
    if locked and locked.get("has_data"):
        row = section_header(ws, row, "Locked-in NII")
        row = write_header_row(ws, row, ["metric", "value"])
        ws.cell(row=row, column=1, value="Total NII (12M)")
        ws.cell(row=row, column=2, value=float(locked.get("total_nii", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        ws.cell(row=row, column=1, value="Locked NII (Fixed-rate)")
        ws.cell(row=row, column=2, value=float(locked.get("locked_nii", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        ws.cell(row=row, column=1, value="Locked %")
        ws.cell(row=row, column=2, value=float(locked.get("locked_pct", 0) or 0) / 100.0).number_format = FMT_PERCENT
        row += 1
        ws.cell(row=row, column=1, value="Certainty-floor narrative valid?")
        ws.cell(row=row, column=2, value="Yes" if locked.get("pct_meaningful", True) else "No (Total NII ≤ 0)")
        row += 1
        by_ccy = locked.get("by_currency") or {}
        if by_ccy:
            row += 1
            ws.cell(row=row, column=1, value="Per Currency").font = Font(bold=True)
            row += 1
            row = write_header_row(ws, row, ["currency", "total_nii", "locked_nii", "locked_pct", "meaningful"])
            for ccy, v in by_ccy.items():
                ws.cell(row=row, column=1, value=ccy)
                ws.cell(row=row, column=2, value=float(v.get("total_nii", 0) or 0)).number_format = FMT_CURRENCY
                ws.cell(row=row, column=3, value=float(v.get("locked_nii", 0) or 0)).number_format = FMT_CURRENCY
                ws.cell(row=row, column=4, value=float(v.get("locked_pct", 0) or 0) / 100.0).number_format = FMT_PERCENT
                ws.cell(row=row, column=5, value="Yes" if v.get("pct_meaningful", True) else "No")
                row += 1

    set_column_widths(ws, [34, 22, 22, 22, 10, 16])
    freeze(ws, f"A{data_start}")
    footer(ws, date_run=date_run, source_key="summary.kpis + summary.coc_ytd + summary.locked_in_nii")


def _write_coc(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("CoC Decomposition")
    coc = data.get("coc", {}) or {}
    table = coc.get("table") if coc.get("has_data") else None
    if not table:
        write_empty_state(ws, "CoC data not available for this run.")
        return

    df = pd.DataFrame(table)
    # Ensure a predictable column order
    preferred = [
        "month", "GrossCarry",
        "FundingCost_Simple", "PnL_Simple", "FundingRate_Simple",
        "FundingCost_Compounded", "PnL_Compounded", "FundingRate_Compounded",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    df = df[cols]

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    style_header_row(ws, row=1, n_cols=len(cols))

    # Apply formats
    fmt_map = {
        "GrossCarry": FMT_CURRENCY,
        "FundingCost_Simple": FMT_CURRENCY,
        "PnL_Simple": FMT_CURRENCY,
        "FundingCost_Compounded": FMT_CURRENCY,
        "PnL_Compounded": FMT_CURRENCY,
        "FundingRate_Simple": FMT_PERCENT,
        "FundingRate_Compounded": FMT_PERCENT,
    }
    apply_format_by_header(ws, fmt_map)

    # Color-scale conditional formatting on PnL columns
    for col_name in ("PnL_Simple", "PnL_Compounded", "GrossCarry"):
        if col_name in cols:
            add_color_scale(ws, cols.index(col_name) + 1, start_row=2, end_row=ws.max_row)

    autofit_columns(ws)
    freeze(ws, "B2")
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="coc.table")


def _write_deal_pnl(wb: Workbook, data: dict, deals: Optional[pd.DataFrame], date_run: str) -> None:
    ws = wb.create_sheet("Deal P&L")
    deal_df = data.get("pnl_by_deal_df")
    if deal_df is None or deal_df.empty:
        write_empty_state(ws, "Deal-level P&L not available.")
        return

    # Enrich with bank-native taxonomy when present
    if deals is not None and {"Dealid", "IAS Book", "Category2"}.issubset(deals.columns):
        tax = (
            deals[["Dealid", "IAS Book", "Category2"]]
            .drop_duplicates(subset=["Dealid"])
            .assign(Dealid=lambda d: d["Dealid"].astype(str))
        )
        deal_df = deal_df.assign(Dealid=lambda d: d["Dealid"].astype(str)).merge(
            tax, on="Dealid", how="left"
        )

    # Focus the P&L report on the base shock; sort by |PnL_Simple| desc
    if "Shock" in deal_df.columns:
        deal_df = deal_df[deal_df["Shock"].astype(str) == "0"].copy()

    # Drop rows where every economic column is zero — the full 60-month grid
    # emits zero-filled placeholders for dormant months, which buries the real
    # signal under tens of thousands of empty rows. Keep any row with a
    # non-trivial nominal, carry, or P&L so the sheet stays informative.
    econ_cols = [c for c in ("Nominal", "GrossCarry", "FundingCost_Simple", "PnL_Simple")
                 if c in deal_df.columns]
    if econ_cols:
        mask = False
        for c in econ_cols:
            mask = mask | (deal_df[c].abs().fillna(0) > 1e-6)
        deal_df = deal_df[mask].copy()

    if "PnL_Simple" in deal_df.columns:
        deal_df["_abs"] = deal_df["PnL_Simple"].abs().fillna(0)
        deal_df = deal_df.sort_values("_abs", ascending=False).drop(columns=["_abs"])

    # openpyxl cannot serialize pd.Period; coerce to strings
    if "Month" in deal_df.columns:
        deal_df["Month"] = deal_df["Month"].astype(str)

    export_cols = [c for c in [
        "Dealid", "Counterparty", "Currency", "Product", "Direction",
        "Périmètre TOTAL", "IAS Book", "Category2", "Month",
        "Nominal", "Amount", "Maturitydate", "is_floating",
        "Clientrate", "OISfwd", "RateRef",
        "GrossCarry", "FundingCost_Simple", "PnL_Simple",
        "FundingRate_Simple",
        "FundingCost_Compounded", "PnL_Compounded",
        "FundingRate_Compounded",
    ] if c in deal_df.columns]
    deal_df = deal_df[export_cols]

    for r in dataframe_to_rows(deal_df, index=False, header=True):
        ws.append(r)
    style_header_row(ws, row=1, n_cols=len(export_cols))

    fmt_map = {
        "Nominal": FMT_CURRENCY,
        "Amount": FMT_CURRENCY,
        "GrossCarry": FMT_CURRENCY,
        "FundingCost_Simple": FMT_CURRENCY,
        "PnL_Simple": FMT_CURRENCY,
        "FundingCost_Compounded": FMT_CURRENCY,
        "PnL_Compounded": FMT_CURRENCY,
        "Clientrate": FMT_PERCENT,
        "OISfwd": FMT_PERCENT,
        "RateRef": FMT_PERCENT,
        "FundingRate_Simple": FMT_PERCENT,
        "FundingRate_Compounded": FMT_PERCENT,
        "Maturitydate": FMT_DATE,
    }
    apply_format_by_header(ws, fmt_map)

    # Column widths — keys narrower, numeric slightly wider
    widths = []
    for h in export_cols:
        if h in {"Dealid", "Currency", "Product", "Direction"}:
            widths.append(12)
        elif h in {"Counterparty", "Périmètre TOTAL", "IAS Book", "Category2", "Month"}:
            widths.append(18)
        elif h == "is_floating":
            widths.append(12)
        else:
            widths.append(20)
    set_column_widths(ws, widths)

    freeze(ws, "B2")
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="pnl_by_deal_df (filtered to shock=0)")


def _write_dod_bridge(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Day-over-Day Bridge")
    summary = data.get("summary", {}) or {}
    bridge = summary.get("dod_bridge")
    if not bridge:
        write_empty_state(ws, "No prior run available — day-over-day bridge skipped.")
        return

    row = title_block(ws, "Day-over-Day P&L Bridge (Base shock)",
                      "Compares current run's base-shock P&L to the prior run, per currency.",
                      span_cols=4)

    data_start = write_header_row(ws, row, ["Currency", "Previous", "Current", "Delta"])
    row = data_start

    for entry in bridge:
        ccy = entry.get("currency", "")
        prev = float(entry.get("previous", 0) or 0)
        curr = float(entry.get("current", 0) or 0)
        delta = float(entry.get("delta", curr - prev) or 0)
        ws.cell(row=row, column=1, value=ccy)
        ws.cell(row=row, column=2, value=prev).number_format = FMT_CURRENCY
        ws.cell(row=row, column=3, value=curr).number_format = FMT_CURRENCY
        dc = ws.cell(row=row, column=4, value=delta)
        dc.number_format = FMT_CURRENCY
        apply_sign_fill(dc, delta)
        if str(ccy).lower() == "total":
            for c in range(1, 5):
                ws.cell(row=row, column=c).font = Font(bold=True)
        row += 1

    set_column_widths(ws, [14, 22, 22, 22])
    freeze(ws, f"A{data_start}")
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="summary.dod_bridge")


def _render_attribution_block(
    ws, attribution: dict, title: str, subtitle: str, date_run: str, source_key: str,
) -> None:
    """Core P&L Explain rendering — summary, by-currency, new/matured deal lists.

    Shared between the Day-over-Day (``Attribution Waterfall``) and
    Month-over-Month (``Attribution MoM``) sheets.
    """
    row = title_block(ws, title, subtitle, span_cols=3)

    summary = attribution.get("summary") or {}
    row = write_header_row(ws, row, ["Driver", "Value"])
    for label, key in (
        ("Previous NII", "prev_nii"),
        ("Rate Effect", "rate_effect"),
        ("Spread Effect", "spread_effect"),
        ("Time / Roll-down", "time_effect"),
        ("New Deals", "new_deal_effect"),
        ("Matured Deals", "matured_deal_effect"),
        ("Current NII", "curr_nii"),
        ("Delta (Current − Previous)", "delta"),
    ):
        val = summary.get(key)
        if val is None:
            continue
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=float(val))
        cell.number_format = FMT_CURRENCY
        if key == "delta":
            ws.cell(row=row, column=1).font = Font(bold=True)
            cell.font = Font(bold=True)
            apply_sign_fill(cell, val)
        row += 1

    counts = [(k, summary.get(k)) for k in ("n_existing", "n_new", "n_matured")]
    if any(v is not None for _, v in counts):
        row += 1
        ws.cell(row=row, column=1, value="Deal counts").font = Font(bold=True)
        row += 1
        for label, key in (("Existing", "n_existing"), ("New", "n_new"), ("Matured", "n_matured")):
            v = summary.get(key)
            if v is None:
                continue
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=int(v)).number_format = FMT_INT
            row += 1

    by_ccy = attribution.get("by_currency") or {}
    if by_ccy:
        row += 2
        row = section_header(ws, row, "Per-currency breakdown")
        row = write_header_row(ws, row, ["Currency", "Previous NII", "Current NII", "Delta", "Rate Effect"])
        for ccy, entry in by_ccy.items():
            prev = float(entry.get("prev_nii", entry.get("existing_prev_pnl", 0)) or 0)
            curr = float(entry.get("curr_nii", entry.get("existing_curr_pnl", 0)) or 0)
            delta = float(entry.get("delta", curr - prev) or 0)
            rate_fx = float(entry.get("rate_effect", 0) or 0)
            ws.cell(row=row, column=1, value=ccy)
            ws.cell(row=row, column=2, value=prev).number_format = FMT_CURRENCY
            ws.cell(row=row, column=3, value=curr).number_format = FMT_CURRENCY
            dc = ws.cell(row=row, column=4, value=delta)
            dc.number_format = FMT_CURRENCY
            apply_sign_fill(dc, delta)
            ws.cell(row=row, column=5, value=rate_fx).number_format = FMT_CURRENCY
            row += 1

    TOP_N = 15
    new_deals = attribution.get("new_deals") or []
    if new_deals:
        row += 2
        row = section_header(
            ws, row, f"Top new deals (by |P&L|) — showing {min(TOP_N, len(new_deals))} of {len(new_deals)}",
        )
        row = write_header_row(
            ws, row, ["Deal ID", "Counterparty", "CCY", "Product", "Nominal", "P&L contribution"],
        )
        for d in new_deals[:TOP_N]:
            ws.cell(row=row, column=1, value=str(d.get("deal_id", "")))
            ws.cell(row=row, column=2, value=str(d.get("counterparty", "")))
            ws.cell(row=row, column=3, value=str(d.get("currency", "")))
            ws.cell(row=row, column=4, value=str(d.get("product", "")))
            ws.cell(row=row, column=5, value=float(d.get("nominal", 0) or 0)).number_format = FMT_CURRENCY
            pv = float(d.get("pnl", 0) or 0)
            pc = ws.cell(row=row, column=6, value=pv)
            pc.number_format = FMT_CURRENCY
            apply_sign_fill(pc, pv)
            row += 1

    matured_deals = attribution.get("matured_deals") or []
    if matured_deals:
        row += 2
        row = section_header(
            ws, row, f"Top matured deals (by |P&L lost|) — showing {min(TOP_N, len(matured_deals))} of {len(matured_deals)}",
        )
        row = write_header_row(
            ws, row, ["Deal ID", "Counterparty", "CCY", "Product", "Nominal", "P&L lost"],
        )
        for d in matured_deals[:TOP_N]:
            ws.cell(row=row, column=1, value=str(d.get("deal_id", "")))
            ws.cell(row=row, column=2, value=str(d.get("counterparty", "")))
            ws.cell(row=row, column=3, value=str(d.get("currency", "")))
            ws.cell(row=row, column=4, value=str(d.get("product", "")))
            ws.cell(row=row, column=5, value=float(d.get("nominal", 0) or 0)).number_format = FMT_CURRENCY
            # pnl_lost is the negated prev P&L — a positive value means NII lost to maturity
            lv = float(d.get("pnl_lost", 0) or 0)
            lc = ws.cell(row=row, column=6, value=lv)
            lc.number_format = FMT_CURRENCY
            apply_sign_fill(lc, -lv)  # invert: loss of income shown in red
            row += 1

    set_column_widths(ws, [32, 22, 22, 22, 22, 24])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key=source_key)


def _write_attribution(wb: Workbook, data: dict, date_run: str) -> None:
    """Day-over-Day P&L Explain — driven by ``dashboard_data['attribution']``."""
    ws = wb.create_sheet("Attribution Waterfall")
    attribution = data.get("attribution") or {}
    if not attribution.get("has_data"):
        write_empty_state(ws, "Attribution requires a prior run — skipped.")
        return
    _render_attribution_block(
        ws, attribution,
        title="P&L Attribution Waterfall — Day-over-Day",
        subtitle="Decomposes \u0394NII (current vs prior run) into drivers.",
        date_run=date_run, source_key="attribution",
    )


def _write_attribution_mom(wb: Workbook, data: dict, date_run: str) -> None:
    """Month-over-Month P&L Explain — driven by ``dashboard_data['attribution_mom']``.

    Anchor is the latest explain snapshot on-or-before the last day of the
    previous month. When no such snapshot exists (first run of the month with
    no prior April run), the sheet shows an empty-state message explaining how
    the anchor accumulates.
    """
    ws = wb.create_sheet("P&L Explain (MoM)")
    mom = data.get("attribution_mom") or {}
    if not mom.get("has_data"):
        write_empty_state(
            ws,
            "MoM P&L Explain unavailable — no explain snapshot exists on or before "
            "the last day of the previous month. The anchor is populated by running "
            "`render-pnl --format pnl-xlsx` on that day (or any earlier day of the "
            "prior month, whichever was last).",
        )
        return

    summary = mom.get("summary") or {}
    prev_d = summary.get("prev_date", "\u2014")
    curr_d = summary.get("curr_date", date_run)
    _render_attribution_block(
        ws, mom,
        title=f"P&L Explain — Month-over-Month ({prev_d} \u2192 {curr_d})",
        subtitle=(
            "Decomposes \u0394NII vs the last explain snapshot on-or-before month-start. "
            "Same decomposition as DoD (rate / spread / time / new / matured)."
        ),
        date_run=date_run, source_key="attribution_mom",
    )


def _write_monthly_series(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Monthly Series")
    series = data.get("pnl_series") or {}
    if not series.get("has_data"):
        write_empty_state(ws, "Monthly P&L series not available.")
        return

    months = series.get("months") or []
    by_ccy = series.get("by_currency") or {}
    rates_month_raw = series.get("date_rates_month", "")
    # months come pre-formatted (e.g. "Apr-26"); rates_month_raw is "2026-04".
    # Normalize to the same label for the highlight comparison; tolerate
    # malformed/empty input by falling back to the raw string.
    if rates_month_raw:
        try:
            rates_month = pd.Period(rates_month_raw, freq="M").strftime("%b-%y")
        except (ValueError, TypeError):
            rates_month = rates_month_raw
    else:
        rates_month = ""

    row = title_block(ws, "Monthly P&L Series",
                      f"Rows = month; columns = currency × shock. "
                      f"Realized vs forecast split at {rates_month}.",
                      span_cols=max(4, len(by_ccy) * 3 + 1))

    shock_keys = ["shock_0", "shock_50", "shock_wirp"]
    shock_labels = {"shock_0": "Base", "shock_50": "+50bp", "shock_wirp": "WIRP"}

    currencies = sorted(by_ccy.keys())
    col_headers = ["Month"]
    col_specs: list[tuple[str, str]] = []  # (ccy, shock_key)
    for ccy in currencies:
        shocks_for_ccy = by_ccy.get(ccy, {})
        for sk in shock_keys:
            if sk in shocks_for_ccy:
                col_headers.append(f"{ccy} {shock_labels[sk]}")
                col_specs.append((ccy, sk))

    header_row = row
    row = write_header_row(ws, row, col_headers)

    # Each by_currency[ccy][shock] contains a list parallel to months
    for m_idx, month in enumerate(months):
        ws.cell(row=row, column=1, value=str(month))
        for col_idx, (ccy, sk) in enumerate(col_specs, start=2):
            vals = by_ccy.get(ccy, {}).get(sk) or []
            if m_idx < len(vals):
                v = vals[m_idx]
                if v is not None:
                    ws.cell(row=row, column=col_idx, value=float(v)).number_format = FMT_CURRENCY
        # Highlight the rates-month row as the realized/forecast boundary
        if rates_month and str(month) == rates_month:
            for c in range(1, len(col_headers) + 1):
                ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFE699")
        row += 1

    set_column_widths(ws, [12] + [16] * (len(col_headers) - 1))
    # Color-scale heat-map across the numeric grid (all currency × shock columns).
    data_start = header_row + 1
    data_end = row - 1
    if len(col_headers) > 1 and data_end >= data_start:
        first_col = get_column_letter(2)
        last_col = get_column_letter(len(col_headers))
        add_color_scale_range(
            ws, f"{first_col}{data_start}:{last_col}{data_end}",
            diverging_at_zero=True,
        )
    freeze(ws, f"B{header_row + 1}")
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="pnl_series")


def _write_daily_projection(wb: Workbook, data: dict, date_run: str) -> None:
    """Daily P&L projection from dateRun through month-end for Central (shock=0) and WIRP.

    Two blocks side-by-side: one per shock. Rows = calendar days; columns per
    currency + Total. Friday rows carry Sat+Sun accrual (d_i=3) by ISDA 2021
    convention, so summing down a currency column gives the period NII.
    """
    ws = wb.create_sheet("Daily Projection (MTD)")
    proj = data.get("daily_projection") or {}
    if not proj.get("has_data"):
        write_empty_state(
            ws,
            "Daily projection unavailable — requires live engine matrices (--format pnl-xlsx/all).",
        )
        return

    start_d = proj.get("start_date", "")
    end_d = proj.get("end_date", "")
    blocks = [
        ("Central (Base / shock=0)", proj.get("central")),
        ("WIRP (Market-implied)", proj.get("wirp")),
    ]
    blocks = [(label, df) for label, df in blocks if isinstance(df, pd.DataFrame) and not df.empty]
    if not blocks:
        write_empty_state(ws, "Daily projection dataframes are empty.")
        return

    # Union of currencies across blocks (stable order: CHF, EUR, USD, GBP, then alpha-rest)
    preferred = ["CHF", "EUR", "USD", "GBP"]
    all_ccy: list[str] = []
    for _, df in blocks:
        for c in df["Currency"].astype(str).unique().tolist():
            if c not in all_ccy:
                all_ccy.append(c)
    ordered_ccy = [c for c in preferred if c in all_ccy] + sorted(c for c in all_ccy if c not in preferred)

    total_cols_per_block = 1 + len(ordered_ccy) + 1  # Date + currencies + Total
    span = total_cols_per_block * len(blocks) + (len(blocks) - 1)  # separator cols
    row = title_block(
        ws,
        f"Daily P&L Projection — {start_d} → {end_d}",
        f"Per-day NII accrual by currency (Fri carries Sat+Sun). "
        f"Side-by-side: Central vs WIRP.",
        span_cols=max(span, 6),
    )

    # Two side-by-side blocks separated by a blank column
    block_starts: list[tuple[int, str, pd.DataFrame]] = []
    col_cursor = 1
    for label, df in blocks:
        block_starts.append((col_cursor, label, df))
        col_cursor += total_cols_per_block + 1  # +1 separator

    # Block title row
    title_row = row
    for col_start, label, _df in block_starts:
        cell = ws.cell(row=title_row, column=col_start, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(
            start_row=title_row, start_column=col_start,
            end_row=title_row, end_column=col_start + total_cols_per_block - 1,
        )
    row += 1

    # Header row (Date | CCYs... | Total) per block
    header_row = row
    for col_start, _label, _df in block_starts:
        ws.cell(row=header_row, column=col_start, value="Date")
        for i, ccy in enumerate(ordered_ccy, start=1):
            ws.cell(row=header_row, column=col_start + i, value=ccy)
        ws.cell(row=header_row, column=col_start + len(ordered_ccy) + 1, value="Total")
    style_header_row(ws, header_row, n_cols=col_cursor - 2)
    row += 1

    # Collect the union of dates across blocks (preserve chronological order)
    all_dates: list[pd.Timestamp] = []
    for _, _lbl, df in block_starts:
        for d in pd.to_datetime(df["Date"]).unique():
            if d not in all_dates:
                all_dates.append(d)
    all_dates = sorted(all_dates)

    data_start = row
    for d in all_dates:
        for col_start, _label, df in block_starts:
            dc = ws.cell(row=row, column=col_start, value=d)
            dc.number_format = FMT_DATE
            sub = df[pd.to_datetime(df["Date"]) == d]
            total = 0.0
            any_val = False
            for i, ccy in enumerate(ordered_ccy, start=1):
                mask = sub["Currency"].astype(str) == ccy
                if mask.any():
                    v = float(sub.loc[mask, "PnL_Daily"].sum())
                    ws.cell(row=row, column=col_start + i, value=v).number_format = FMT_CURRENCY
                    total += v
                    any_val = True
            if any_val:
                tc = ws.cell(row=row, column=col_start + len(ordered_ccy) + 1, value=total)
                tc.number_format = FMT_CURRENCY
                tc.font = Font(bold=True)
                apply_sign_fill(tc, total)
        row += 1
    data_end = row - 1

    # Totals row per block
    for col_start, _label, df in block_starts:
        ws.cell(row=row, column=col_start, value="Total").font = Font(bold=True)
        grand_total = 0.0
        for i, ccy in enumerate(ordered_ccy, start=1):
            mask = df["Currency"].astype(str) == ccy
            if mask.any():
                v = float(df.loc[mask, "PnL_Daily"].sum())
                tc = ws.cell(row=row, column=col_start + i, value=v)
                tc.number_format = FMT_CURRENCY
                tc.font = Font(bold=True)
                apply_sign_fill(tc, v)
                grand_total += v
        gc = ws.cell(row=row, column=col_start + len(ordered_ccy) + 1, value=grand_total)
        gc.number_format = FMT_CURRENCY
        gc.font = Font(bold=True)
        apply_sign_fill(gc, grand_total)
    row += 1

    # Heat-map over per-currency numeric cells (exclude Date + Total) per block
    if data_end >= data_start:
        for col_start, _label, _df in block_starts:
            first = get_column_letter(col_start + 1)
            last = get_column_letter(col_start + len(ordered_ccy))
            add_color_scale_range(
                ws, f"{first}{data_start}:{last}{data_end}",
                diverging_at_zero=True,
            )

    # Column widths: 12 for Date, 15 per currency+Total, 2 separator
    widths: list[int] = []
    for b_idx, _bs in enumerate(block_starts):
        widths.append(12)  # Date
        widths.extend([15] * len(ordered_ccy))
        widths.append(16)  # Total
        if b_idx < len(block_starts) - 1:
            widths.append(2)  # separator
    set_column_widths(ws, widths)
    freeze(ws, f"A{header_row + 1}")
    footer(ws, date_run=date_run, source_key="daily_projection (central + wirp)")


def _write_realized_daily_pnl(wb: Workbook, data: dict, date_run: str) -> None:
    """Realized daily P&L from bank export — BOOK1 accrual/IAS + BOOK2 MTM.

    Unlike the forecast projection (engine recomputed), this sheet reports the
    numbers the bank itself booked on ``date_run``. Rows = currencies; columns
    = one per book bucket (BOOK1 Accrual, BOOK1 IAS, BOOK2 MTM) + row Total.
    """
    ws = wb.create_sheet("Realized Daily P&L")
    realized = data.get("realized_daily_pnl") or {}
    if not realized.get("has_data"):
        write_empty_state(
            ws,
            "Realized daily P&L unavailable — requires bank-native input "
            "(Book1 with PnL_Acc_Adj / PnL_Realized, Book2 non-IRS with PnL_Realized).",
        )
        return

    pivot = realized.get("pivot")
    if pivot is None or not isinstance(pivot, pd.DataFrame) or pivot.empty:
        write_empty_state(ws, "Realized daily P&L pivot is empty.")
        return

    book_cols = [c for c in pivot.columns if c != "currency"]
    headers = ["Currency"] + book_cols + ["Total"]

    row = title_block(
        ws,
        f"Realized Daily P&L — {realized.get('date_run', '')}",
        "Bank-reported daily P&L (not forecast). BOOK1 = accrual IAS/ORC, BOOK2 = MTM.",
        span_cols=len(headers),
    )
    header_row = row
    row = write_header_row(ws, row, headers)

    col_totals = {col: 0.0 for col in book_cols}
    for _, r in pivot.iterrows():
        ccy = str(r["currency"])
        ws.cell(row=row, column=1, value=ccy).font = Font(bold=True)
        row_total = 0.0
        for i, col in enumerate(book_cols, start=2):
            v = float(r[col] or 0.0)
            c = ws.cell(row=row, column=i, value=v)
            c.number_format = FMT_CURRENCY
            apply_sign_fill(c, v)
            row_total += v
            col_totals[col] += v
        tc = ws.cell(row=row, column=len(headers), value=row_total)
        tc.number_format = FMT_CURRENCY
        tc.font = Font(bold=True)
        apply_sign_fill(tc, row_total)
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    grand_total = 0.0
    for i, col in enumerate(book_cols, start=2):
        v = col_totals[col]
        tc = ws.cell(row=row, column=i, value=v)
        tc.number_format = FMT_CURRENCY
        tc.font = Font(bold=True)
        apply_sign_fill(tc, v)
        grand_total += v
    gc = ws.cell(row=row, column=len(headers), value=grand_total)
    gc.number_format = FMT_CURRENCY
    gc.font = Font(bold=True)
    apply_sign_fill(gc, grand_total)

    set_column_widths(ws, [12] + [22] * len(book_cols) + [18])
    freeze(ws, f"A{header_row + 1}")
    footer(ws, date_run=date_run, source_key="realized_daily_pnl")


def _write_book2_delta(
    wb: Workbook, sheet_name: str, title: str, subtitle_template: str,
    block: dict, date_run: str, source_key: str,
) -> None:
    """Shared layout for BOOK2 ΔMTM sheets (DoD / MTD).

    ``subtitle_template`` may contain ``{prev_date}``, which is replaced by the
    resolved prior snapshot date (or "—" when absent).
    """
    ws = wb.create_sheet(sheet_name)
    if not block or not block.get("has_data"):
        write_empty_state(
            ws,
            f"{title} unavailable — requires a prior strategy snapshot and today's Book2 MTM.",
        )
        return

    prev_date = block.get("prev_date") or "\u2014"
    row = title_block(
        ws, title,
        subtitle_template.format(prev_date=prev_date),
        span_cols=5,
    )
    header_row = row
    row = write_header_row(ws, row, ["Currency", "#Deals", "MTM Prev", "MTM Today", "\u0394MTM"])

    rows = block.get("rows") or []
    for r in rows:
        ws.cell(row=row, column=1, value=str(r.get("currency", ""))).font = Font(bold=True)
        ws.cell(row=row, column=2, value=int(r.get("n_deals", 0) or 0)).number_format = FMT_INT
        ws.cell(row=row, column=3, value=float(r.get("mtm_prev", 0.0))).number_format = FMT_CURRENCY
        ws.cell(row=row, column=4, value=float(r.get("mtm_today", 0.0))).number_format = FMT_CURRENCY
        delta = float(r.get("delta", 0.0))
        dc = ws.cell(row=row, column=5, value=delta)
        dc.number_format = FMT_CURRENCY
        dc.font = Font(bold=True)
        apply_sign_fill(dc, delta)
        row += 1

    totals = block.get("totals") or {}
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    n_total = sum(int(r.get("n_deals", 0) or 0) for r in rows)
    ws.cell(row=row, column=2, value=n_total).number_format = FMT_INT
    ws.cell(row=row, column=3, value=float(totals.get("mtm_prev", 0.0))).number_format = FMT_CURRENCY
    ws.cell(row=row, column=4, value=float(totals.get("mtm_today", 0.0))).number_format = FMT_CURRENCY
    total_delta = float(totals.get("delta", 0.0))
    tc = ws.cell(row=row, column=5, value=total_delta)
    tc.number_format = FMT_CURRENCY
    tc.font = Font(bold=True)
    apply_sign_fill(tc, total_delta)

    set_column_widths(ws, [12, 10, 20, 20, 20])
    freeze(ws, f"A{header_row + 1}")
    footer(ws, date_run=date_run, source_key=source_key)


def _write_book2_delta_dod(wb: Workbook, data: dict, date_run: str) -> None:
    """BOOK2 ΔMTM — Day-over-Day (vs last snapshot strictly before today)."""
    _write_book2_delta(
        wb,
        sheet_name="BOOK2 \u0394MTM - Day-over-Day",
        title="BOOK2 \u0394MTM — Day-over-Day",
        subtitle_template="MTM(today) \u2212 MTM(prev). Prior snapshot: {prev_date}.",
        block=data.get("book2_delta_dod") or {},
        date_run=date_run,
        source_key="book2_delta_dod",
    )


def _write_book2_delta_mtd(wb: Workbook, data: dict, date_run: str) -> None:
    """BOOK2 ΔMTM — Month-to-Date (vs last snapshot on-or-before month start)."""
    _write_book2_delta(
        wb,
        sheet_name="BOOK2 \u0394MTM - MTD",
        title="BOOK2 \u0394MTM — Month-to-Date",
        subtitle_template="MTM(today) \u2212 MTM(month-start). Anchor snapshot: {prev_date}.",
        block=data.get("book2_delta_mtd") or {},
        date_run=date_run,
        source_key="book2_delta_mtd",
    )


def _write_book1_realized_mtd(wb: Workbook, data: dict, date_run: str) -> None:
    """BOOK1 Realized — MTD (sum of daily bank-reported BOOK1 P&L for the month)."""
    ws = wb.create_sheet("BOOK1 Realized - MTD")
    block = data.get("book1_realized_mtd") or {}
    if not block or not block.get("has_data"):
        days = int(block.get("days_counted", 0) or 0) if block else 0
        msg = (
            f"BOOK1 Realized MTD unavailable — {days} daily snapshot(s) found but no "
            "non-zero BOOK1 figures. MTD accumulates from kpi_snapshots/*_realized.json "
            "(one file per pnl-xlsx run)."
        )
        write_empty_state(ws, msg)
        return

    month_start = block.get("month_start", "")
    date_end = block.get("date_run", "")
    days = int(block.get("days_counted", 0) or 0)

    row = title_block(
        ws,
        f"BOOK1 Realized — MTD ({month_start} \u2192 {date_end})",
        f"Sum of bank-reported BOOK1 daily P&L across {days} snapshot(s). "
        "Accrual = @PnL_Acc_Estim_Adj; IAS = [Daily] PnL IAS - ORC.",
        span_cols=5,
    )
    header_row = row
    row = write_header_row(
        ws, row, ["Currency", "BOOK1 Accrual MTD", "BOOK1 IAS MTD", "BOOK1 Total MTD", "# Deals"],
    )

    rows = block.get("rows") or []
    for r in rows:
        ws.cell(row=row, column=1, value=str(r.get("currency", ""))).font = Font(bold=True)
        accr = float(r.get("book1_accrual", 0.0))
        ias = float(r.get("book1_ias", 0.0))
        total = float(r.get("book1_total", accr + ias))
        ac = ws.cell(row=row, column=2, value=accr)
        ac.number_format = FMT_CURRENCY
        apply_sign_fill(ac, accr)
        ic = ws.cell(row=row, column=3, value=ias)
        ic.number_format = FMT_CURRENCY
        apply_sign_fill(ic, ias)
        tc = ws.cell(row=row, column=4, value=total)
        tc.number_format = FMT_CURRENCY
        tc.font = Font(bold=True)
        apply_sign_fill(tc, total)
        # Column 5 left blank — deal count is not tracked in snapshot; kept for future use.
        row += 1

    totals = block.get("totals") or {}
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    accr_t = float(totals.get("book1_accrual", 0.0))
    ias_t = float(totals.get("book1_ias", 0.0))
    tot_t = float(totals.get("book1_total", accr_t + ias_t))
    for col, v in ((2, accr_t), (3, ias_t), (4, tot_t)):
        c = ws.cell(row=row, column=col, value=v)
        c.number_format = FMT_CURRENCY
        c.font = Font(bold=True)
        apply_sign_fill(c, v)

    set_column_widths(ws, [12, 22, 22, 22, 10])
    freeze(ws, f"A{header_row + 1}")
    footer(ws, date_run=date_run, source_key="book1_realized_mtd")


def _write_top_contributors(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Top Contributors")
    summary = data.get("summary", {}) or {}
    top5 = summary.get("top5") or []

    row = title_block(ws, "Top Contributors & Detractors",
                      "Top 5 by |P&L| + deal-level contributors / detractors (12-month horizon).",
                      span_cols=4)

    row = section_header(ws, row, "Top 5 by Absolute P&L (summary.top5)")
    if top5:
        row = write_header_row(ws, row, ["Currency", "Product", "P&L"])
        for entry in top5:
            ws.cell(row=row, column=1, value=entry.get("currency", ""))
            ws.cell(row=row, column=2, value=entry.get("product", ""))
            pnl = float(entry.get("pnl", 0) or 0)
            cell = ws.cell(row=row, column=3, value=pnl)
            cell.number_format = FMT_CURRENCY
            apply_sign_fill(cell, pnl)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Not available.")
        row += 1

    deal_df = data.get("pnl_by_deal_df")
    if deal_df is not None and not deal_df.empty and "PnL_Simple" in deal_df.columns:
        df = deal_df
        if "Shock" in df.columns:
            df = df[df["Shock"].astype(str) == "0"]
        # Match the Top 5 window: first 12 months of the grid. Without this, Top
        # 10 / Bottom 10 aggregate across 60 months while Top 5 reflects only 12,
        # producing cross-panel rankings that don't align.
        if "Month" in df.columns:
            months_12 = sorted(df["Month"].astype(str).unique())[:12]
            df = df[df["Month"].astype(str).isin(months_12)]
        # pnl_by_deal_df is per (deal, month); aggregate to per-deal totals.
        key_cols = [c for c in ["Dealid", "Counterparty", "Currency", "Product", "Direction"]
                    if c in df.columns]
        if key_cols:
            agg = df.groupby(key_cols, dropna=False)["PnL_Simple"].sum().reset_index()
            agg = agg.sort_values("PnL_Simple", ascending=False)

            for title, frame in (
                ("Top 10 Deals (contributors, base shock, 12M)", agg.head(10)),
                ("Bottom 10 Deals (detractors, base shock, 12M)", agg.tail(10).iloc[::-1]),
            ):
                row += 2
                row = section_header(ws, row, title)
                row = write_header_row(ws, row, key_cols + ["PnL_Simple"])
                for _, r in frame.iterrows():
                    for c, h in enumerate(key_cols, start=1):
                        ws.cell(row=row, column=c, value=r[h])
                    pnl = float(r["PnL_Simple"] or 0)
                    pc = ws.cell(row=row, column=len(key_cols) + 1, value=pnl)
                    pc.number_format = FMT_CURRENCY
                    apply_sign_fill(pc, pnl)
                    row += 1

    set_column_widths(ws, [18, 20, 14, 14, 12, 22])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="summary.top5 + pnl_by_deal_df")


def _write_strategy(wb: Workbook, data: dict, date_run: str) -> None:
    """Strategy IAS 4-leg P&L decomposition (IAM/LD-NHCD, IAM/LD-HCD, BND-NHCD, BND-HCD)."""
    ws = wb.create_sheet("Strategy P&L")
    strategy = data.get("strategy") or {}
    if not strategy.get("has_data"):
        write_empty_state(
            ws,
            "No Strategy IAS decomposition — requires deals with IAS hedge designation.",
        )
        return

    row = title_block(ws, "P&L by Strategy IAS Leg",
                      "4-leg decomposition (IAM/LD-NHCD, IAM/LD-HCD, BND-NHCD, BND-HCD). Base shock.",
                      span_cols=6)

    # --- Leg summary table ---
    table = strategy.get("table") or []
    if table:
        row = section_header(ws, row, "Leg Summary")
        row = write_header_row(
            ws, row,
            ["Leg", "Currency", "Direction", "Total P&L", "Avg Nominal", "Avg RateRef", "Avg OIS"],
        )
        total_pnl = 0.0
        for entry in table:
            ws.cell(row=row, column=1, value=entry.get("leg", ""))
            ws.cell(row=row, column=2, value=entry.get("currency", ""))
            ws.cell(row=row, column=3, value=entry.get("direction", ""))
            pnl = float(entry.get("pnl", 0) or 0)
            pc = ws.cell(row=row, column=4, value=pnl)
            pc.number_format = FMT_CURRENCY
            apply_sign_fill(pc, pnl)
            ws.cell(row=row, column=5, value=float(entry.get("nominal", 0) or 0)).number_format = FMT_CURRENCY
            ws.cell(row=row, column=6, value=float(entry.get("rate_ref", 0) or 0)).number_format = FMT_PERCENT
            ws.cell(row=row, column=7, value=float(entry.get("ois_fwd", 0) or 0)).number_format = FMT_PERCENT
            total_pnl += pnl
            row += 1

        # Total row
        ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
        tc = ws.cell(row=row, column=4, value=total_pnl)
        tc.number_format = FMT_CURRENCY
        tc.font = Font(bold=True)
        apply_sign_fill(tc, total_pnl)
        row += 1

    # --- Monthly series per leg ---
    months = strategy.get("months") or []
    legs = strategy.get("legs") or {}
    if months and legs:
        row += 2
        row = section_header(ws, row, "Monthly P&L per Leg")
        leg_names = sorted(legs.keys())
        header_row = row
        row = write_header_row(ws, row, ["Month"] + leg_names + ["Total"])
        data_start = row

        for m_idx, month in enumerate(months):
            ws.cell(row=row, column=1, value=str(month))
            month_total = 0.0
            for c, leg in enumerate(leg_names, start=2):
                vals = legs.get(leg, {}).get("values") or []
                if m_idx < len(vals):
                    v = float(vals[m_idx] or 0)
                    ws.cell(row=row, column=c, value=v).number_format = FMT_CURRENCY
                    month_total += v
            tc = ws.cell(row=row, column=len(leg_names) + 2, value=month_total)
            tc.number_format = FMT_CURRENCY
            apply_sign_fill(tc, month_total)
            row += 1

        data_end = row - 1
        # Color scale across per-leg cells only (exclude Month + Total)
        if data_end >= data_start and len(leg_names) > 0:
            first_col = get_column_letter(2)
            last_col = get_column_letter(len(leg_names) + 1)
            add_color_scale_range(
                ws, f"{first_col}{data_start}:{last_col}{data_end}",
                diverging_at_zero=True,
            )
        freeze(ws, f"B{header_row + 1}")

    set_column_widths(ws, [14, 18, 14, 20, 20, 14, 14])
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="strategy")


def _write_hedge_effectiveness(wb: Workbook, data: dict, date_run: str) -> None:
    """Strategy IAS cross-book consolidated view — IAS 39 / IFRS 9 corridor test."""
    ws = wb.create_sheet("Hedge Effectiveness")
    sc = data.get("strategy_consolidated") or {}
    if not sc.get("has_data"):
        write_empty_state(
            ws,
            "No hedge effectiveness data — requires deals with Strategy IAS designation "
            "and both hedged items and hedging instruments.",
        )
        return

    row = title_block(
        ws,
        "Hedge Effectiveness — Strategy IAS Consolidated",
        "Corridor [80%, 125%]: Effectiveness = -\u0394MtM / \u0394FV. \u0394 vs prior snapshot.",
        span_cols=11,
    )

    summary = sc.get("summary") or {}
    row = section_header(ws, row, "Summary")
    row = write_header_row(ws, row, ["Total", "OK", "Under", "Over", "N/A", "Multi-CCY"])
    ws.cell(row=row, column=1, value=int(summary.get("n_total", 0) or 0)).number_format = FMT_INT
    ok_cell = ws.cell(row=row, column=2, value=int(summary.get("n_ok", 0) or 0))
    ok_cell.number_format = FMT_INT
    if summary.get("n_ok"):
        ok_cell.fill = PatternFill("solid", fgColor=FILL_POS)
    under_cell = ws.cell(row=row, column=3, value=int(summary.get("n_under", 0) or 0))
    under_cell.number_format = FMT_INT
    if summary.get("n_under"):
        under_cell.fill = PatternFill("solid", fgColor=FILL_NEG)
    over_cell = ws.cell(row=row, column=4, value=int(summary.get("n_over", 0) or 0))
    over_cell.number_format = FMT_INT
    if summary.get("n_over"):
        over_cell.fill = PatternFill("solid", fgColor=FILL_NEG)
    ws.cell(row=row, column=5, value=int(summary.get("n_na", 0) or 0)).number_format = FMT_INT
    ws.cell(row=row, column=6, value=int(summary.get("n_multi_ccy", 0) or 0)).number_format = FMT_INT
    row += 2

    rows = sc.get("rows") or []
    if rows:
        row = section_header(ws, row, "Relationships")
        header_row = row
        row = write_header_row(
            ws, row,
            [
                "Strategy IAS", "Type", "CCY", "#Hedged", "#Hedging",
                "Hedged FV", "\u0394FV", "IRS MtM", "\u0394MtM",
                "Effectiveness", "Corridor",
            ],
        )
        for r in rows:
            ws.cell(row=row, column=1, value=r.get("strategy_ias", ""))
            ws.cell(row=row, column=2, value=r.get("hedge_type", ""))
            ws.cell(row=row, column=3, value=r.get("currencies", ""))
            ws.cell(row=row, column=4, value=int(r.get("n_hedged", 0) or 0)).number_format = FMT_INT
            ws.cell(row=row, column=5, value=int(r.get("n_hedging", 0) or 0)).number_format = FMT_INT

            fv = r.get("hedged_clean_fv_today")
            if fv is not None:
                ws.cell(row=row, column=6, value=float(fv)).number_format = FMT_CURRENCY
            dfv = r.get("hedged_clean_dFV")
            if dfv is not None:
                dfv_val = float(dfv)
                dc = ws.cell(row=row, column=7, value=dfv_val)
                dc.number_format = FMT_CURRENCY
                apply_sign_fill(dc, dfv_val)

            mtm = r.get("hedging_irs_mtm_today")
            if mtm is not None:
                ws.cell(row=row, column=8, value=float(mtm)).number_format = FMT_CURRENCY
            dmtm = r.get("hedging_irs_dMtM")
            if dmtm is not None:
                dmtm_val = float(dmtm)
                mc = ws.cell(row=row, column=9, value=dmtm_val)
                mc.number_format = FMT_CURRENCY
                apply_sign_fill(mc, dmtm_val)

            eff = r.get("effectiveness_ratio")
            if eff is not None:
                ws.cell(row=row, column=10, value=float(eff)).number_format = FMT_PERCENT

            flag = (r.get("corridor_flag") or "").lower()
            flag_label = {
                "ok": "OK",
                "under": "UNDER",
                "over": "OVER",
                "multi_ccy": "MULTI-CCY",
                "na": "N/A",
            }.get(flag, flag.upper() or "N/A")
            fc = ws.cell(row=row, column=11, value=flag_label)
            if flag == "ok":
                fc.fill = PatternFill("solid", fgColor=FILL_POS)
                fc.font = Font(bold=True)
            elif flag in ("under", "over"):
                fc.fill = PatternFill("solid", fgColor=FILL_NEG)
                fc.font = Font(bold=True)
            row += 1

        freeze(ws, f"A{header_row + 2}")
        add_autofilter(ws)

    set_column_widths(ws, [16, 10, 10, 10, 10, 18, 18, 18, 18, 14, 12])
    footer(ws, date_run=date_run, source_key="strategy_consolidated")


def _write_ftp(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("3-Way Margin (FTP)")
    ftp = data.get("ftp") or {}
    if not ftp.get("has_data"):
        coverage = ftp.get("coverage") or {}
        msg = "FTP data not available."
        if coverage:
            msg += f" Coverage: {coverage.get('coverage_pct', 0)}%"
        write_empty_state(ws, msg)
        return

    row = title_block(ws, "3-Way Margin Decomposition (FTP)",
                      "Total NII = Client Margin (ClientRate − FTP) + ALM Margin (FTP − OIS).",
                      span_cols=6)

    # By perimeter
    perimeters = ftp.get("perimeters") or {}
    if perimeters:
        row = section_header(ws, row, "By Perimeter (CC / WM / CIB)")
        row = write_header_row(ws, row, [
            "Perimeter", "Client Margin", "ALM Margin", "Total NII",
            "Deal Count", "Avg Client bps", "Avg ALM bps",
        ])
        for peri, entry in perimeters.items():
            ws.cell(row=row, column=1, value=peri)
            ws.cell(row=row, column=2, value=float(entry.get("client_margin", 0) or 0)).number_format = FMT_CURRENCY
            ws.cell(row=row, column=3, value=float(entry.get("alm_margin", 0) or 0)).number_format = FMT_CURRENCY
            total = float(entry.get("total_nii", 0) or 0)
            tc = ws.cell(row=row, column=4, value=total)
            tc.number_format = FMT_CURRENCY
            apply_sign_fill(tc, total)
            ws.cell(row=row, column=5, value=int(entry.get("deal_count", 0) or 0)).number_format = FMT_INT
            ws.cell(row=row, column=6, value=float(entry.get("avg_client_margin_bps", 0) or 0))
            ws.cell(row=row, column=7, value=float(entry.get("avg_alm_margin_bps", 0) or 0))
            row += 1

    # By currency
    by_currency = ftp.get("by_currency") or {}
    if by_currency:
        row += 2
        row = section_header(ws, row, "By Currency")
        row = write_header_row(ws, row, ["Currency", "Client Margin", "ALM Margin", "Total NII", "Deal Count"])
        for ccy, entry in by_currency.items():
            ws.cell(row=row, column=1, value=ccy)
            ws.cell(row=row, column=2, value=float(entry.get("client_margin", 0) or 0)).number_format = FMT_CURRENCY
            ws.cell(row=row, column=3, value=float(entry.get("alm_margin", 0) or 0)).number_format = FMT_CURRENCY
            total = float(entry.get("total_nii", 0) or 0)
            tc = ws.cell(row=row, column=4, value=total)
            tc.number_format = FMT_CURRENCY
            apply_sign_fill(tc, total)
            ws.cell(row=row, column=5, value=int(entry.get("deal_count", 0) or 0)).number_format = FMT_INT
            row += 1

    # Top 10 deals by FTP margin
    top_deals = ftp.get("top_deals") or []
    if top_deals:
        row += 2
        row = section_header(ws, row, "Top 10 Deals by |ALM Margin|")
        # Column keys are inferred from the first record
        keys = list(top_deals[0].keys())
        row = write_header_row(ws, row, keys)
        for td in top_deals:
            for c, k in enumerate(keys, start=1):
                v = td.get(k)
                cell = ws.cell(row=row, column=c, value=v)
                if isinstance(v, (int, float)) and "margin" in k.lower():
                    cell.number_format = '#,##0.0' if "bps" in k.lower() else FMT_CURRENCY
            row += 1

    # Coverage
    coverage = ftp.get("coverage") or {}
    if coverage:
        row += 2
        ws.cell(row=row, column=1, value="FTP Coverage").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value="Deals with FTP")
        ws.cell(row=row, column=2, value=int(coverage.get("with_ftp", 0) or 0)).number_format = FMT_INT
        row += 1
        ws.cell(row=row, column=1, value="Deals without FTP")
        ws.cell(row=row, column=2, value=int(coverage.get("without_ftp", 0) or 0)).number_format = FMT_INT
        row += 1
        ws.cell(row=row, column=1, value="Coverage %")
        ws.cell(row=row, column=2, value=float(coverage.get("coverage_pct", 0) or 0) / 100.0).number_format = FMT_PERCENT
        row += 1

    set_column_widths(ws, [18, 20, 20, 22, 14, 16, 16])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="ftp")


def _write_fixed_float(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Fixed vs Floating")
    ff = data.get("fixed_float") or {}
    if not ff.get("has_data"):
        write_empty_state(ws, "Fixed/Floating breakdown not available.")
        return

    row = title_block(ws, "Fixed vs Floating Mix",
                      "Portfolio notional split between fixed-rate and floating-rate deals.",
                      span_cols=5)

    # Mix block
    mix = ff.get("mix") or {}
    if mix:
        row = section_header(ws, row, "Overall Mix")
        row = write_header_row(ws, row, ["Type", "Count", "Nominal", "Share"])
        for t in ("Fixed", "Floating"):
            entry = mix.get(t)
            if not entry:
                continue
            ws.cell(row=row, column=1, value=t)
            ws.cell(row=row, column=2, value=int(entry.get("count", 0) or 0)).number_format = FMT_INT
            ws.cell(row=row, column=3, value=float(entry.get("nominal", 0) or 0)).number_format = FMT_CURRENCY
            ws.cell(row=row, column=4, value=round(float(entry.get("pct", 0) or 0) / 100.0, 4)).number_format = FMT_PERCENT
            row += 1

    # By currency
    by_ccy = ff.get("by_currency") or {}
    if by_ccy:
        row += 2
        row = section_header(ws, row, "By Currency")
        row = write_header_row(ws, row, ["Currency", "Fixed", "Floating", "Fixed %", "Floating %", "# Fixed", "# Floating"])
        for ccy, entry in by_ccy.items():
            ws.cell(row=row, column=1, value=ccy)
            ws.cell(row=row, column=2, value=float(entry.get("fixed", 0) or 0)).number_format = FMT_CURRENCY
            ws.cell(row=row, column=3, value=float(entry.get("floating", 0) or 0)).number_format = FMT_CURRENCY
            # Rounding after the /100 conversion: the builder already rounds to 1
            # decimal place of percent, but float representation of e.g. 54.8/100
            # yields 0.5479999… in the formula bar. Re-rounding to 4 decimals
            # (= 2 decimals of percent) matches FMT_PERCENT and cleans the raw.
            ws.cell(row=row, column=4, value=round(float(entry.get("fixed_pct", 0) or 0) / 100.0, 4)).number_format = FMT_PERCENT
            ws.cell(row=row, column=5, value=round(float(entry.get("floating_pct", 0) or 0) / 100.0, 4)).number_format = FMT_PERCENT
            ws.cell(row=row, column=6, value=int(entry.get("count_fixed", 0) or 0)).number_format = FMT_INT
            ws.cell(row=row, column=7, value=int(entry.get("count_floating", 0) or 0)).number_format = FMT_INT
            row += 1

    # Sensitivity
    sens = ff.get("sensitivity") or {}
    if sens:
        row += 2
        row = section_header(ws, row, "Rate Sensitivity (+50bp vs Base)")
        for k, v in sens.items():
            ws.cell(row=row, column=1, value=k)
            if isinstance(v, (int, float)):
                cell = ws.cell(row=row, column=2, value=float(v))
                cell.number_format = FMT_CURRENCY if "delta" in k.lower() or "pnl" in k.lower() else '#,##0.00'
            else:
                ws.cell(row=row, column=2, value=str(v))
            row += 1

    set_column_widths(ws, [16, 20, 20, 14, 14, 14, 14])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="fixed_float")


def _write_budget(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Budget vs Actual")
    budget = data.get("budget") or {}
    if not budget.get("has_data"):
        write_empty_state(ws, "Budget data not loaded — supply budget.xlsx to populate.")
        return

    row = title_block(ws, "Budget vs Actual",
                      "Monthly actuals vs board-approved budget per currency, plus YTD variance.",
                      span_cols=6)

    # YTD block first
    ytd = budget.get("ytd") or {}
    if ytd:
        row = section_header(ws, row, "YTD Variance")
        row = write_header_row(ws, row, ["Metric", "Value"])
        ws.cell(row=row, column=1, value="Actual")
        ws.cell(row=row, column=2, value=float(ytd.get("actual", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        ws.cell(row=row, column=1, value="Budget")
        ws.cell(row=row, column=2, value=float(ytd.get("budget", 0) or 0)).number_format = FMT_CURRENCY
        row += 1
        var = float(ytd.get("variance", 0) or 0)
        ws.cell(row=row, column=1, value="Variance")
        vc = ws.cell(row=row, column=2, value=var)
        vc.number_format = FMT_CURRENCY
        apply_sign_fill(vc, var)
        row += 1
        ws.cell(row=row, column=1, value="Variance %")
        ws.cell(row=row, column=2, value=float(ytd.get("variance_pct", 0) or 0) / 100.0).number_format = FMT_PERCENT
        row += 1

    # Variance waterfall
    waterfall = budget.get("variance_waterfall") or []
    if waterfall:
        row += 2
        row = section_header(ws, row, "Variance Waterfall (YTD)")
        row = write_header_row(ws, row, ["Step", "Value", "Type"])
        for step in waterfall:
            ws.cell(row=row, column=1, value=step.get("label", ""))
            val = float(step.get("value", 0) or 0)
            cell = ws.cell(row=row, column=2, value=val)
            cell.number_format = FMT_CURRENCY
            stype = step.get("type", "")
            if stype in ("base", "total"):
                for c in range(1, 4):
                    ws.cell(row=row, column=c).font = Font(bold=True)
            elif stype == "effect":
                apply_sign_fill(cell, val)
            ws.cell(row=row, column=3, value=stype)
            row += 1

    # Monthly table per currency
    months = budget.get("months") or []
    by_ccy = budget.get("by_currency") or {}
    if months and by_ccy:
        row += 2
        row = section_header(ws, row, "Monthly Actual vs Budget")
        row = write_header_row(ws, row, ["Month", "Currency", "Actual", "Budget", "Variance"])
        for ccy, entry in by_ccy.items():
            actuals = entry.get("actual", []) or []
            budgets = entry.get("budget", []) or []
            variances = entry.get("variance", []) or []
            for i, m in enumerate(months):
                ws.cell(row=row, column=1, value=str(m))
                ws.cell(row=row, column=2, value=ccy)
                if i < len(actuals):
                    ws.cell(row=row, column=3, value=float(actuals[i] or 0)).number_format = FMT_CURRENCY
                if i < len(budgets):
                    ws.cell(row=row, column=4, value=float(budgets[i] or 0)).number_format = FMT_CURRENCY
                if i < len(variances):
                    v = float(variances[i] or 0)
                    vc = ws.cell(row=row, column=5, value=v)
                    vc.number_format = FMT_CURRENCY
                    apply_sign_fill(vc, v)
                row += 1

    set_column_widths(ws, [16, 16, 22, 22, 22])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="budget")


def _write_forecast_tracking(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Forecast Tracking")
    tracking = data.get("forecast_tracking") or {}
    if not tracking.get("has_data"):
        write_empty_state(ws, "Forecast history not loaded — run backfill to accumulate.")
        return

    row = title_block(ws, "Forecast Tracking",
                      "Evolution of the 12M forecast over prior runs; drift/stability stats.",
                      span_cols=6)

    # Stats block
    stats = tracking.get("stats") or {}
    if stats:
        row = section_header(ws, row, "Stability Statistics")
        row = write_header_row(ws, row, ["Metric", "Value"])
        for k, v in stats.items():
            ws.cell(row=row, column=1, value=k.replace("_", " ").title())
            if isinstance(v, (int, float)):
                cell = ws.cell(row=row, column=2, value=float(v))
                low = k.lower()
                if "pct" in low or "stability" in low:
                    cell.number_format = FMT_PERCENT
                elif "bps" in low:
                    cell.number_format = '#,##0.0" bps"'
                else:
                    cell.number_format = FMT_CURRENCY
            else:
                ws.cell(row=row, column=2, value=str(v))
            row += 1

    # Total series
    dates = tracking.get("dates") or []
    total = tracking.get("total") or []
    if dates and total:
        row += 2
        row = section_header(ws, row, "Total 12M Forecast Over Time")
        row = write_header_row(ws, row, ["Run Date", "Total NII Forecast"])
        for i, d in enumerate(dates):
            ws.cell(row=row, column=1, value=str(d))
            if i < len(total):
                v = total[i]
                if v is not None:
                    ws.cell(row=row, column=2, value=float(v)).number_format = FMT_CURRENCY
            row += 1

    # Per-currency series
    by_ccy = tracking.get("by_currency") or {}
    if dates and by_ccy:
        row += 2
        row = section_header(ws, row, "Per-Currency Evolution")
        ccy_list = sorted(by_ccy.keys())
        row = write_header_row(ws, row, ["Run Date"] + ccy_list)
        for i, d in enumerate(dates):
            ws.cell(row=row, column=1, value=str(d))
            for c, ccy in enumerate(ccy_list, start=2):
                series = by_ccy.get(ccy) or []
                if i < len(series) and series[i] is not None:
                    ws.cell(row=row, column=c, value=float(series[i])).number_format = FMT_CURRENCY
            row += 1

    # Revisions table (largest forecast-to-forecast deltas)
    revisions = tracking.get("revisions") or []
    if revisions:
        row += 2
        row = section_header(ws, row, "Notable Revisions")
        keys = list(revisions[0].keys())
        row = write_header_row(ws, row, keys)
        for rev in revisions:
            for c, k in enumerate(keys, start=1):
                v = rev.get(k)
                cell = ws.cell(row=row, column=c, value=v)
                if isinstance(v, (int, float)) and any(t in k.lower() for t in ["delta", "revision", "value", "nii"]):
                    cell.number_format = FMT_CURRENCY
            row += 1

    set_column_widths(ws, [16, 22, 22, 22, 22, 22])
    freeze(ws, "A5")
    footer(ws, date_run=date_run, source_key="forecast_tracking")


def _write_alerts(wb: Workbook, data: dict, date_run: str) -> None:
    ws = wb.create_sheet("Alerts")
    alerts_data = data.get("pnl_alerts") or {}
    alerts = alerts_data.get("alerts") or []

    row = title_block(ws, "P&L Alerts",
                      "Threshold-based alerts on moves, shocks, breaches.",
                      span_cols=7)

    summary = alerts_data.get("summary") or {}
    if summary:
        row = section_header(ws, row, "Severity Counts")
        row = write_header_row(ws, row, ["Severity", "Count"])
        for sev in ("critical", "high", "medium", "low"):
            count = summary.get(sev)
            if count is None:
                continue
            ws.cell(row=row, column=1, value=sev.title())
            ws.cell(row=row, column=2, value=int(count)).number_format = FMT_INT
            if count:
                ws.cell(row=row, column=1).fill = PatternFill(
                    "solid", fgColor=SEVERITY_BADGE_FILLS[sev],
                )
            row += 1

    if alerts:
        row += 2
        row = section_header(ws, row, "Alert Details")
        preferred = ["type", "severity", "metric", "current", "threshold", "message", "recommendation"]
        keys_set: set[str] = set()
        for a in alerts:
            keys_set.update(a.keys())
        keys = [k for k in preferred if k in keys_set] + sorted(k for k in keys_set if k not in preferred)
        row = write_header_row(ws, row, keys)
        for a in alerts:
            sev = str(a.get("severity", "")).lower()
            row_color = SEVERITY_ROW_FILLS.get(sev)
            fill = PatternFill("solid", fgColor=row_color) if row_color else None
            for c, k in enumerate(keys, start=1):
                v = a.get(k)
                cell = ws.cell(row=row, column=c, value=v)
                if isinstance(v, (int, float)) and k in ("current", "threshold"):
                    cell.number_format = '#,##0.00'
                if fill:
                    cell.fill = fill
            row += 1
    else:
        row += 2
        write_empty_state(ws, "No active alerts.", row=row)

    set_column_widths(ws, [14, 12, 20, 16, 16, 50, 50])
    freeze(ws, "A5")
    add_autofilter(ws)
    footer(ws, date_run=date_run, source_key="pnl_alerts")


def _fill_cover_jump_links(wb: Workbook) -> None:
    """Populate Cover's 'Jump to sheet' block with hyperlinks to every sheet."""
    if "Cover" not in wb.sheetnames:
        return
    cover = wb["Cover"]
    anchor = getattr(cover, "_cover_jump_links_anchor", None)
    if anchor is None:
        return
    skip = {"Cover"}
    col1_names = [s for s in wb.sheetnames if s not in skip]
    # Two-column layout for compactness
    per_col = (len(col1_names) + 1) // 2
    for i, name in enumerate(col1_names):
        row = anchor + (i % per_col)
        col = 1 if i < per_col else 3
        write_jump_link(cover, row, col, name, name)


def _write_diagnostics(
    wb: Workbook,
    diagnostics: list[tuple[str, str]],
    date_run: str,
    data: dict | None = None,
    deals: Optional[pd.DataFrame] = None,
) -> None:
    ws = wb.create_sheet("Diagnostics")

    row = title_block(ws, "Diagnostics", f"Per-sheet build status for run {date_run}.", span_cols=2)

    # Per-sheet status
    row = write_header_row(ws, row, ["Sheet", "Status"])
    ok_count = 0
    for name, status in diagnostics:
        ws.cell(row=row, column=1, value=name)
        c = ws.cell(row=row, column=2, value=status)
        c.fill = PatternFill("solid", fgColor=FILL_POS if status == "OK" else FILL_NEG)
        if status == "OK":
            ok_count += 1
        row += 1

    # Run-level metadata
    if data is not None:
        row += 2
        row = section_header(ws, row, "Run Metadata")
        row = write_header_row(ws, row, ["Metric", "Value"])

        meta_rows: list[tuple[str, object]] = [("Run date", date_run)]

        # Deal counts
        deal_df = data.get("pnl_by_deal_df")
        if deal_df is not None and hasattr(deal_df, "empty") and not deal_df.empty:
            meta_rows.append(("Deal P&L rows", len(deal_df)))
            if "Dealid" in deal_df.columns:
                meta_rows.append(("Unique deals", int(deal_df["Dealid"].astype(str).nunique())))
            if "Shock" in deal_df.columns:
                shocks = sorted(str(s) for s in deal_df["Shock"].unique())
                meta_rows.append(("Shocks", ", ".join(shocks)))
        if deals is not None and hasattr(deals, "empty") and not deals.empty:
            meta_rows.append(("Deals in portfolio", len(deals)))
            if "Currency" in deals.columns:
                meta_rows.append(("Currencies", ", ".join(sorted(deals["Currency"].dropna().astype(str).unique()))))
            if "FTP" in deals.columns:
                n_ftp = int(deals["FTP"].notna().sum())
                meta_rows.append(("Deals with FTP", f"{n_ftp} / {len(deals)}"))

        # Feature availability flags
        meta_rows.append(("Prior-run comparison", "Yes" if data.get("attribution", {}).get("has_data") else "No"))
        meta_rows.append(("Budget loaded", "Yes" if data.get("budget", {}).get("has_data") else "No"))
        meta_rows.append(("Forecast history", "Yes" if data.get("forecast_tracking", {}).get("has_data") else "No"))
        meta_rows.append(("FTP data", "Yes" if data.get("ftp", {}).get("has_data") else "No"))
        meta_rows.append(("Alerts count", len((data.get("pnl_alerts") or {}).get("alerts", []))))
        meta_rows.append(("Sheets written (OK)", f"{ok_count} / {len(diagnostics)}"))

        for label, value in meta_rows:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if isinstance(value, int) and not isinstance(value, bool):
                cell.number_format = FMT_INT
            row += 1

    set_column_widths(ws, [32, 60])
    freeze(ws, "A4")
