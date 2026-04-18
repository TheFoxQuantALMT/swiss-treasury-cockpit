"""Excel export for the P&L dashboard data.

Exports dashboard data to a multi-sheet Excel workbook using openpyxl.
Each dashboard tab becomes a worksheet with formatted tables.

Each sheet is written through a per-section helper wrapped in its own
try/except so a single bad section does not cascade. A trailing Diagnostics
sheet records the outcome, and a post-write pass applies treasury styling
(header formatting, number formats, freeze panes, autofilter, conditional
color-scales, print layout) without changing sheet names or column schemas.
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill

from cockpit.export.styles import (
    FILL_NEG,
    FILL_POS,
    FMT_CURRENCY,
    FMT_INT,
    FMT_PERCENT,
    SEVERITY_ROW_FILLS,
    add_autofilter,
    add_color_scale,
    apply_format_by_header,
    autofit_columns,
    freeze,
    hide_gridlines,
    set_column_widths,
    setup_print_layout,
    style_header_row,
    title_block,
)


# Columns that should be formatted as currency if present as a header
_CURRENCY_HEADERS = {
    "total", "realized", "forecast",
    "gross_carry", "funding_cost", "pnl_simple", "pnl_compounded",
    "GrossCarry", "FundingCost_Simple", "PnL_Simple",
    "FundingCost_Compounded", "PnL_Compounded",
    "Nominal", "Amount",
    "eve", "EVE", "Value",
    "actual", "budget", "variance",
    "client_margin", "alm_margin", "total_nii",
    "current", "threshold",
}

_PERCENT_HEADERS = {
    "realized_pct", "duration",
    "FundingRate_Simple", "FundingRate_Compounded",
    "Clientrate", "OISfwd", "RateRef",
    "utilization_pct", "variance_pct",
}

_DATE_HEADERS = {"Maturitydate", "month", "Month"}

_PNL_COL_FOR_COLOR_SCALE = {
    "PnL_Simple", "PnL_Compounded", "pnl_simple", "pnl_compounded",
    "variance", "Value", "total", "eve",
}


def export_dashboard_to_excel(
    dashboard_data: dict,
    output_path: Path | str,
    date_run: str = "",
    deals: "pd.DataFrame | None" = None,
) -> Path | None:
    """Export dashboard data dict to Excel workbook.

    Creates sheets for: Summary, Sensitivity, EVE, Alerts, Limits, FTP, plus a
    bank-native ``Synthesis`` sheet (rolled up by IAS Book × @Category2) when
    the deals frame carries that taxonomy.

    Args:
        dashboard_data: Dict from build_pnl_dashboard_data().
        output_path: Output .xlsx path.
        date_run: Date string for metadata.
        deals: Optional deals DataFrame carrying Dealid/IAS Book/Category2
            (bank-native input). Enables the Synthesis sheet and enriches
            Deal PnL with the taxonomy columns.

    Returns:
        Path to generated file, or None on failure.
    """
    output_path = Path(output_path)
    diagnostics: list[tuple[str, str]] = []

    def _run(name: str, fn: Callable[[], None]) -> None:
        try:
            fn()
            diagnostics.append((name, "OK"))
        except Exception as e:  # noqa: BLE001 — per-sheet isolation is intentional
            diagnostics.append((name, f"SKIPPED: {e}"))
            print(f"[excel-export] Sheet '{name}' skipped: {e}")

    try:
        with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
            _run("Summary", lambda: _write_summary(writer, dashboard_data))
            _run("PnL Simple vs Compounded",
                 lambda: _write_pnl_simple_compounded(writer, dashboard_data))
            _run("Sensitivity", lambda: _write_sensitivity(writer, dashboard_data))
            _run("EVE", lambda: _write_eve(writer, dashboard_data))
            _run("Alerts", lambda: _write_alerts(writer, dashboard_data))
            _run("Limits", lambda: _write_limits(writer, dashboard_data))
            _run("CoC", lambda: _write_coc(writer, dashboard_data))
            _run("CoC by Currency", lambda: _write_coc_by_currency(writer, dashboard_data))
            _run("FTP", lambda: _write_ftp(writer, dashboard_data))
            _run("Deal PnL", lambda: _write_deal_pnl(writer, dashboard_data, deals))
            _run("Synthesis",
                 lambda: _write_synthesis(writer, dashboard_data, deals))

            # Metadata always writes — tests rely on this invariant.
            _write_metadata(writer, date_run)
            diagnostics.append(("Metadata", "OK"))

            # Diagnostics + polish pass after all sheets exist.
            _write_diagnostics(writer, diagnostics, date_run)
            _polish_workbook(writer, date_run)

        return output_path
    except Exception as e:  # noqa: BLE001
        print(f"[excel-export] Failed: {e}")
        return None


# ---------------------------------------------------------------------------
# Per-section writers — behaviour matches the pre-refactor inline code.
# ---------------------------------------------------------------------------

def _write_summary(writer: pd.ExcelWriter, data: dict) -> None:
    summary = data.get("summary", {}) or {}
    kpis = summary.get("kpis", {}) or {}
    if not kpis:
        return
    rows = []
    for shock, entry in kpis.items():
        if isinstance(entry, dict):
            rows.append({"shock": shock, **entry})
    if not rows:
        return
    pd.DataFrame(rows).to_excel(writer, sheet_name="Summary", index=False)


def _write_pnl_simple_compounded(writer: pd.ExcelWriter, data: dict) -> None:
    summary = data.get("summary", {}) or {}
    coc_ytd = summary.get("coc_ytd")
    if not coc_ytd:
        return
    pd.DataFrame([coc_ytd]).to_excel(
        writer, sheet_name="PnL Simple vs Compounded", index=False,
    )


def _write_sensitivity(writer: pd.ExcelWriter, data: dict) -> None:
    sensitivity = data.get("sensitivity", {}) or {}
    rows = sensitivity.get("rows")
    if not rows:
        return
    pd.DataFrame(rows).to_excel(writer, sheet_name="Sensitivity", index=False)


def _write_eve(writer: pd.ExcelWriter, data: dict) -> None:
    eve = data.get("eve", {}) or {}
    if not (eve.get("has_data") and eve.get("by_currency")):
        return
    rows = [{"currency": k, **v} for k, v in eve["by_currency"].items()]
    pd.DataFrame(rows).to_excel(writer, sheet_name="EVE", index=False)


def _write_alerts(writer: pd.ExcelWriter, data: dict) -> None:
    alerts = data.get("pnl_alerts", {}) or {}
    items = alerts.get("alerts")
    if not items:
        return
    pd.DataFrame(items).to_excel(writer, sheet_name="Alerts", index=False)


def _write_limits(writer: pd.ExcelWriter, data: dict) -> None:
    limits = data.get("limits", {}) or {}
    if not (limits.get("has_data") and limits.get("limit_items")):
        return
    pd.DataFrame(limits["limit_items"]).to_excel(
        writer, sheet_name="Limits", index=False,
    )


def _write_coc(writer: pd.ExcelWriter, data: dict) -> None:
    coc = data.get("coc", {}) or {}
    if not (coc.get("has_data") and coc.get("table")):
        return
    pd.DataFrame(coc["table"]).to_excel(writer, sheet_name="CoC", index=False)


def _write_coc_by_currency(writer: pd.ExcelWriter, data: dict) -> None:
    coc = data.get("coc", {}) or {}
    if not (coc.get("has_data") and coc.get("by_currency") and coc.get("months")):
        return
    months = coc["months"]
    rows = []
    for ccy, by_shock in coc["by_currency"].items():
        shock_data = by_shock.get("shock_0", {})
        if not shock_data:
            continue
        for indice, values in shock_data.items():
            for i, m in enumerate(months):
                if i < len(values):
                    rows.append({
                        "Currency": ccy,
                        "Measure": indice,
                        "Month": m,
                        "Value": values[i],
                    })
    if not rows:
        return
    pd.DataFrame(rows).to_excel(writer, sheet_name="CoC by Currency", index=False)


def _write_ftp(writer: pd.ExcelWriter, data: dict) -> None:
    ftp = data.get("ftp", {}) or {}
    if not (ftp.get("has_data") and ftp.get("by_perimeter")):
        return
    pd.DataFrame(ftp["by_perimeter"]).to_excel(
        writer, sheet_name="FTP", index=False,
    )


def _write_deal_pnl(
    writer: pd.ExcelWriter, data: dict, deals: Optional[pd.DataFrame],
) -> None:
    deal_pnl_raw = data.get("pnl_by_deal_df")
    if deal_pnl_raw is None or deal_pnl_raw.empty:
        return
    has_taxonomy = (
        deals is not None
        and {"Dealid", "IAS Book", "Category2"}.issubset(deals.columns)
    )
    deal_pnl = deal_pnl_raw
    if has_taxonomy:
        tax = (
            deals[["Dealid", "IAS Book", "Category2"]]
            .drop_duplicates(subset=["Dealid"])
            .assign(Dealid=lambda d: d["Dealid"].astype(str))
        )
        deal_pnl = deal_pnl_raw.assign(
            Dealid=lambda d: d["Dealid"].astype(str)
        ).merge(tax, on="Dealid", how="left")
    export_cols = [c for c in [
        "Dealid", "Counterparty", "Currency", "Product", "Direction",
        "Périmètre TOTAL", "IAS Book", "Category2", "Shock", "Month",
        "Nominal", "Amount", "Maturitydate", "is_floating",
        "Clientrate", "OISfwd", "RateRef",
        "GrossCarry", "FundingCost_Simple", "PnL_Simple",
        "FundingRate_Simple",
        "FundingCost_Compounded", "PnL_Compounded",
        "FundingRate_Compounded",
    ] if c in deal_pnl.columns]
    frame = deal_pnl[export_cols].copy()
    if "Month" in frame.columns:
        frame["Month"] = frame["Month"].astype(str)
    frame.to_excel(writer, sheet_name="Deal PnL", index=False)


def _write_synthesis(
    writer: pd.ExcelWriter, data: dict, deals: Optional[pd.DataFrame],
) -> None:
    deal_pnl_raw = data.get("pnl_by_deal_df")
    if deal_pnl_raw is None or deal_pnl_raw.empty or deals is None:
        return
    if not {"Dealid", "IAS Book", "Category2"}.issubset(deals.columns):
        return
    from cockpit.export.synthesis import build_synthesis
    synthesis = build_synthesis(deal_pnl_raw, deals, shock="0")
    if not synthesis.empty:
        synthesis.to_excel(writer, sheet_name="Synthesis", index=False)
    if "Currency" in deals.columns:
        synthesis_ccy = build_synthesis(
            deal_pnl_raw, deals, shock="0", by_currency=True,
        )
        if not synthesis_ccy.empty:
            synthesis_ccy.to_excel(
                writer, sheet_name="Synthesis by Currency", index=False,
            )


def _write_metadata(writer: pd.ExcelWriter, date_run: str) -> None:
    meta = pd.DataFrame([{"date_run": date_run, "export_type": "dashboard"}])
    meta.to_excel(writer, sheet_name="Metadata", index=False)


def _write_diagnostics(
    writer: pd.ExcelWriter, diagnostics: list[tuple[str, str]], date_run: str,
) -> None:
    rows = [{"sheet": name, "status": status} for name, status in diagnostics]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Diagnostics", index=False)


# ---------------------------------------------------------------------------
# Polish pass — applied after all sheets are written.
# ---------------------------------------------------------------------------

def _polish_workbook(writer: pd.ExcelWriter, date_run: str) -> None:
    """Apply header styling, number formats, freeze, autofilter, conditional
    formatting, and print layout to every sheet that exists."""
    wb = writer.book
    for sheet_name in list(writer.sheets.keys()):
        ws = writer.sheets[sheet_name]
        try:
            _polish_sheet(ws, sheet_name, date_run)
        except Exception as e:  # noqa: BLE001
            print(f"[excel-export] Polish skipped for '{sheet_name}': {e}")


def _polish_sheet(ws, sheet_name: str, date_run: str) -> None:
    if ws.max_row < 1:
        return
    # Header row styling — applied if there is at least one row of data below.
    style_header_row(ws, row=1, n_cols=ws.max_column)

    # Number formats by header heuristic
    fmt_map: dict[str, str] = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=c).value
        if not isinstance(header, str):
            continue
        if header in _CURRENCY_HEADERS:
            fmt_map[header] = FMT_CURRENCY
        elif header in _PERCENT_HEADERS:
            fmt_map[header] = FMT_PERCENT
        elif header in _DATE_HEADERS:
            fmt_map[header] = "yyyy-mm-dd"
        elif header in ("deal_count", "count", "count_fixed", "count_floating"):
            fmt_map[header] = FMT_INT
    if fmt_map:
        apply_format_by_header(ws, fmt_map)

    # Conditional color-scale on known P&L-signed columns
    if ws.max_row >= 2:
        for c in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=c).value
            if isinstance(header, str) and header in _PNL_COL_FOR_COLOR_SCALE:
                add_color_scale(ws, c, start_row=2, end_row=ws.max_row)

    # Column widths — autofit for most sheets
    autofit_columns(ws)

    # Freeze + autofilter for tabular sheets; skip for trivially small ones
    if ws.max_row >= 2:
        freeze(ws, "B2")
        add_autofilter(ws)

    # Print layout (all sheets)
    setup_print_layout(ws, date_run=date_run)

    # Hide gridlines on report-style sheets
    if sheet_name in {"Metadata", "Diagnostics"}:
        hide_gridlines(ws)

    # Severity-colored rows on Alerts
    if sheet_name == "Alerts":
        _color_alert_rows(ws)

    # Breach highlighting on Limits
    if sheet_name == "Limits":
        _color_limit_breaches(ws)


def _color_alert_rows(ws) -> None:
    """Colour each alert row based on its severity column."""
    sev_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == "severity":
            sev_col = c
            break
    if sev_col is None:
        return
    for r in range(2, ws.max_row + 1):
        sev = ws.cell(row=r, column=sev_col).value
        if not isinstance(sev, str):
            continue
        fill_color = SEVERITY_ROW_FILLS.get(sev.lower())
        if fill_color:
            fill = PatternFill("solid", fgColor=fill_color)
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill


def _color_limit_breaches(ws) -> None:
    """Highlight limit rows where status is 'red' or utilization > 100%."""
    status_col = None
    util_col = None
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h == "status":
            status_col = c
        elif h == "utilization_pct":
            util_col = c
    if status_col is None and util_col is None:
        return
    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=status_col).value if status_col else None
        util = ws.cell(row=r, column=util_col).value if util_col else None
        is_breach = (status == "red") or (isinstance(util, (int, float)) and util > 100)
        is_warn = (status == "yellow") or (isinstance(util, (int, float)) and 80 <= util <= 100)
        if is_breach:
            fill = PatternFill("solid", fgColor="FCE4D6")
        elif is_warn:
            fill = PatternFill("solid", fgColor="FFE699")
        else:
            continue
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill
