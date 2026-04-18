"""Shared openpyxl formatting helpers for Excel exports.

Centralizes the treasury-grade workbook conventions (colors, number formats,
column widths, header styling) so sheets across exporters render cohesively.
"""
from __future__ import annotations

from typing import Iterable, Optional

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Palette
NAVY = "1F4E79"
WHITE = "FFFFFF"
FILL_POS = "E2EFDA"
FILL_NEG = "FCE4D6"
FILL_BREACH = "C00000"
FILL_WARN = "FFE699"
FILL_MUTED = "F2F2F2"
BORDER_LIGHT = "D9D9D9"

# Severity palettes — kept central so the row tint and the count-badge tint
# don't drift apart. Two variants because alert *rows* read better in pastel
# (matches FILL_POS/FILL_NEG), while *count badges* want strong contrast.
SEVERITY_ROW_FILLS: dict[str, str] = {
    "critical": "FCE4D6",
    "high": "FFE699",
    "medium": "FFF2CC",
    "low": "E2EFDA",
}
SEVERITY_BADGE_FILLS: dict[str, str] = {
    "critical": "C00000",
    "high": "FF8C00",
    "medium": "FFE699",
    "low": "D9E1F2",
}

# Number formats
FMT_CURRENCY = '#,##0;[Red]-#,##0'
FMT_CURRENCY_DECIMALS = '#,##0.00;[Red]-#,##0.00'
FMT_PERCENT = '0.00%;[Red]-0.00%'
FMT_PERCENT_SIGNED = '+0.00%;-0.00%'
FMT_BPS = '#,##0" bps";[Red]-#,##0" bps"'
FMT_INT = '#,##0'
FMT_DATE = 'yyyy-mm-dd'

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_HEADER_FILL = PatternFill("solid", fgColor=NAVY)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_LIGHT),
    right=Side(style="thin", color=BORDER_LIGHT),
    top=Side(style="thin", color=BORDER_LIGHT),
    bottom=Side(style="thin", color=BORDER_LIGHT),
)


def section_header(ws: Worksheet, row: int, text: str, col: int = 1) -> int:
    """Write a navy/bold section header. Returns the next free row."""
    ws.cell(row=row, column=col, value=text).font = Font(bold=True, size=12, color=NAVY)
    return row + 1


def write_header_row(ws: Worksheet, row: int, headers: Iterable[str]) -> int:
    """Write a list of column headers at ``row`` and apply the navy header
    style. Returns the next free row (i.e. ``row + 1``)."""
    headers = list(headers)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row=row, n_cols=len(headers))
    return row + 1


def apply_sign_fill(cell, value: float) -> None:
    """Fill the cell green when value >= 0, red otherwise.

    Also strips ``[Red]`` from the cell's number format if present: the pastel
    fill already communicates sign, and red text on a red fill is unreadable.
    """
    cell.fill = PatternFill("solid", fgColor=FILL_POS if value >= 0 else FILL_NEG)
    fmt = cell.number_format or ""
    if "[Red]" in fmt:
        cell.number_format = fmt.replace("[Red]", "")


def write_empty_state(ws: Worksheet, message: str, row: int = 1, col: int = 1) -> None:
    """Write a single muted-italic message into an otherwise empty sheet."""
    cell = ws.cell(row=row, column=col, value=message)
    cell.font = Font(italic=True, color="595959")


def style_header_row(ws: Worksheet, row: int = 1, n_cols: Optional[int] = None) -> None:
    """Apply navy header styling to the given row across n_cols (or ws.max_column)."""
    end = n_cols if n_cols is not None else ws.max_column
    for c in range(1, end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 24


def set_column_widths(ws: Worksheet, widths: Iterable[int]) -> None:
    """Set column widths from a sequence. None entries are skipped."""
    for i, w in enumerate(widths, start=1):
        if w is not None:
            ws.column_dimensions[get_column_letter(i)].width = w


def autofit_columns(ws: Worksheet, min_w: int = 10, max_w: int = 28, sample: int = 200) -> None:
    """Approximate auto-fit based on a header + first `sample` rows.

    openpyxl cannot measure rendered width; this inspects string lengths as a
    heuristic. Call after data is written.
    """
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row_idx in range(1, min(ws.max_row, sample) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            longest = max(longest, len(str(val)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, longest + 2))


def apply_format(ws: Worksheet, col_idx: int, fmt: str, start_row: int = 2) -> None:
    """Apply a number_format to every cell in a column below the header."""
    for row_idx in range(start_row, ws.max_row + 1):
        ws.cell(row=row_idx, column=col_idx).number_format = fmt


def apply_format_by_header(ws: Worksheet, header_to_fmt: dict[str, str], header_row: int = 1) -> None:
    """Look up each header in the given row and apply its format to that column."""
    headers = {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }
    for name, fmt in header_to_fmt.items():
        col_idx = headers.get(name)
        if col_idx is not None:
            apply_format(ws, col_idx, fmt, start_row=header_row + 1)


def freeze(ws: Worksheet, cell: str = "B2") -> None:
    ws.freeze_panes = cell


def add_autofilter(ws: Worksheet) -> None:
    ws.auto_filter.ref = ws.dimensions


def title_block(ws: Worksheet, title: str, subtitle: Optional[str] = None,
                start_row: int = 1, span_cols: int = 6) -> int:
    """Write a styled title (optional subtitle) block. Returns the next free row."""
    ws.cell(row=start_row, column=1, value=title).font = Font(size=16, bold=True, color=NAVY)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span_cols)
    ws.row_dimensions[start_row].height = 26
    if subtitle:
        ws.cell(row=start_row + 1, column=1, value=subtitle).font = Font(size=10, italic=True, color="595959")
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=span_cols)
        return start_row + 3
    return start_row + 2


def footer(ws: Worksheet, *, date_run: str, source_key: str, row: Optional[int] = None) -> None:
    """Write a muted provenance footer two rows below last data row."""
    target = row if row is not None else ws.max_row + 2
    text = f"Generated: {date_run}   |   Source: dashboard_data[\"{source_key}\"]"
    cell = ws.cell(row=target, column=1, value=text)
    cell.font = Font(size=9, italic=True, color="808080")


def setup_print_layout(ws: Worksheet, *, date_run: str, orientation: str = "landscape",
                       fit_to_width: int = 1) -> None:
    """Apply a consistent print layout: fit-to-width, print header/footer."""
    ws.page_setup.orientation = orientation
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.oddHeader.left.text = "P&&L Daily Report"
    ws.oddHeader.left.size = 9
    ws.oddHeader.right.text = f"Run: {date_run}"
    ws.oddHeader.right.size = 9
    ws.oddFooter.center.text = "Page &P of &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.right.text = "&A"  # sheet name
    ws.oddFooter.right.size = 9


def hide_gridlines(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False


def add_color_scale(ws: Worksheet, col_idx: int, start_row: int = 2,
                    end_row: Optional[int] = None, *,
                    diverging_at_zero: bool = True) -> None:
    """Add a color-scale conditional format on a numeric column.

    When ``diverging_at_zero``, uses a red(neg) → white(0) → green(pos) scale
    anchored at 0. Otherwise a simple red → green min-to-max scale.
    """
    if end_row is None:
        end_row = ws.max_row
    if end_row < start_row:
        return
    letter = get_column_letter(col_idx)
    cell_range = f"{letter}{start_row}:{letter}{end_row}"
    add_color_scale_range(ws, cell_range, diverging_at_zero=diverging_at_zero)


def add_color_scale_range(ws: Worksheet, cell_range: str, *,
                          diverging_at_zero: bool = True) -> None:
    """Same as add_color_scale but takes an A1-style range directly."""
    if diverging_at_zero:
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FCFCFF",
            end_type="max", end_color="63BE7B",
        )
    else:
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            end_type="max", end_color="63BE7B",
        )
    ws.conditional_formatting.add(cell_range, rule)


def write_jump_link(ws: Worksheet, row: int, col: int, label: str, target_sheet: str,
                    target_cell: str = "A1") -> None:
    """Write a clickable hyperlink to another sheet in the same workbook."""
    cell = ws.cell(row=row, column=col, value=label)
    # openpyxl expects sheet names wrapped in single quotes + apostrophe-escaped
    safe_name = target_sheet.replace("'", "''")
    cell.hyperlink = f"#'{safe_name}'!{target_cell}"
    cell.font = Font(color="0563C1", underline="single")
