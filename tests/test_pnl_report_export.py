"""Smoke tests for cockpit.export.pnl_report sheet writers."""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

from cockpit.export.pnl_report import (
    _write_attribution,
    _write_attribution_mom,
    _write_book1_realized_mtd,
    _write_book2_delta_dod,
    _write_book2_delta_mtd,
    _write_daily_projection,
    _write_hedge_effectiveness,
    _write_realized_daily_pnl,
)


def _fresh_wb() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def test_daily_projection_empty_state():
    wb = _fresh_wb()
    _write_daily_projection(wb, {}, "2026-04-18")
    ws = wb["Daily Projection (MTD)"]
    assert ws.max_row >= 1
    assert ws.max_column >= 1


def test_daily_projection_with_data_shapes_side_by_side_blocks():
    dates = pd.date_range("2026-04-14", "2026-04-18", freq="D")
    central = pd.DataFrame([
        {"Date": d, "Currency": ccy, "PnL_Daily": 1000.0 * (1 if ccy == "CHF" else -1)}
        for d in dates for ccy in ("CHF", "EUR")
    ])
    wirp = pd.DataFrame([
        {"Date": d, "Currency": ccy, "PnL_Daily": 1100.0 * (1 if ccy == "CHF" else -1)}
        for d in dates for ccy in ("CHF", "EUR")
    ])
    data = {"daily_projection": {
        "has_data": True, "central": central, "wirp": wirp,
        "start_date": "2026-04-14", "end_date": "2026-04-18",
    }}
    wb = _fresh_wb()
    _write_daily_projection(wb, data, "2026-04-18")
    ws = wb["Daily Projection (MTD)"]

    # Two blocks side-by-side: Date + 2 ccy + Total, separator, Date + 2 ccy + Total
    # Header row contains "Date", "CHF", "EUR", "Total", "", "Date", "CHF", "EUR", "Total"
    header_row = 5  # after title_block (2) + blank (1) + block-title row (1)
    vals = [ws.cell(header_row, c).value for c in range(1, 10)]
    assert "Date" in vals
    assert vals.count("Date") == 2
    assert "CHF" in vals and "EUR" in vals
    assert vals.count("Total") == 2

    # Verify Central total for first date = 1000 + (-1000) = 0
    # Col layout: Date(1), CHF(2), EUR(3), Total(4)
    row6_total = ws.cell(6, 4).value
    assert row6_total == 0.0

    # Grand totals row: sum over 5 days = 0 in central, 0 in wirp (since CHF+EUR offset).
    # Scan rows to find the "Total" label in column A (layout: data rows, total, blank, footer).
    total_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Total"]
    assert len(total_rows) == 1


def test_daily_projection_empty_dfs_falls_back_to_empty_state():
    data = {"daily_projection": {
        "has_data": True,  # flag says yes but frames are empty
        "central": pd.DataFrame(columns=["Date", "Currency", "PnL_Daily"]),
        "wirp": pd.DataFrame(columns=["Date", "Currency", "PnL_Daily"]),
        "start_date": "2026-04-14", "end_date": "2026-04-18",
    }}
    wb = _fresh_wb()
    _write_daily_projection(wb, data, "2026-04-18")
    ws = wb["Daily Projection (MTD)"]
    # Sheet exists; writer should not crash
    assert ws.max_row >= 1


def test_hedge_effectiveness_empty_state():
    wb = _fresh_wb()
    _write_hedge_effectiveness(wb, {}, "2026-04-18")
    assert "Hedge Effectiveness" in wb.sheetnames


def test_hedge_effectiveness_populated():
    data = {"strategy_consolidated": {
        "has_data": True,
        "summary": {"n_total": 2, "n_ok": 1, "n_under": 1, "n_over": 0, "n_na": 0, "n_multi_ccy": 0},
        "rows": [
            {"strategy_ias": "S1", "hedge_type": "FVH", "currencies": "CHF",
             "multi_currency": False, "n_hedged": 1, "n_hedging": 1, "n_hedging_book2": 1,
             "hedged_clean_fv_today": 10_000_000.0, "hedged_clean_dFV": -100_000.0,
             "hedging_irs_mtm_today": 250_000.0, "hedging_irs_dMtM": 99_000.0,
             "effectiveness_ratio": 0.99, "corridor_flag": "ok"},
            {"strategy_ias": "S2", "hedge_type": "FVH", "currencies": "EUR",
             "multi_currency": False, "n_hedged": 1, "n_hedging": 1, "n_hedging_book2": 1,
             "hedged_clean_fv_today": 5_000_000.0, "hedged_clean_dFV": -200_000.0,
             "hedging_irs_mtm_today": 100_000.0, "hedging_irs_dMtM": 100_000.0,
             "effectiveness_ratio": 0.50, "corridor_flag": "under"},
        ],
    }}
    wb = _fresh_wb()
    _write_hedge_effectiveness(wb, data, "2026-04-18")
    ws = wb["Hedge Effectiveness"]
    # Walk for the corridor labels in the last column
    flag_values = {ws.cell(r, 11).value for r in range(1, ws.max_row + 1)}
    assert "OK" in flag_values
    assert "UNDER" in flag_values


def test_realized_daily_pnl_empty_state():
    wb = _fresh_wb()
    _write_realized_daily_pnl(wb, {}, "2026-04-18")
    ws = wb["Realized Daily P&L"]
    assert ws.max_row >= 1


def test_realized_daily_pnl_populated():
    pivot = pd.DataFrame([
        {"currency": "CHF", "BOOK1 Accrual (PnL_Acc_Adj)": 100_000.0,
         "BOOK1 Realized (PnL_IAS)": 20_000.0, "BOOK2 Realized (PnL_MTM)": -5_000.0},
        {"currency": "EUR", "BOOK1 Accrual (PnL_Acc_Adj)": 50_000.0,
         "BOOK1 Realized (PnL_IAS)": 10_000.0, "BOOK2 Realized (PnL_MTM)": 3_000.0},
    ])
    data = {"realized_daily_pnl": {
        "has_data": True,
        "rows": [],  # not consumed by writer
        "pivot": pivot,
        "date_run": "2026-04-18",
    }}
    wb = _fresh_wb()
    _write_realized_daily_pnl(wb, data, "2026-04-18")
    ws = wb["Realized Daily P&L"]

    # Header row has Currency + 3 book columns + Total
    header_row = 4  # title_block spans 2 rows + blank -> header at row 4 (depends on helper)
    # Rather than assert exact row, scan for the row with "Currency" in col A
    hdr_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Currency"]
    assert hdr_rows, "Header row with 'Currency' not found"
    hdr = hdr_rows[0]
    # Total column is last — value should be "Total"
    last_col = ws.max_column
    assert ws.cell(hdr, last_col).value == "Total"

    # Totals row — scan for "Total" label in col A
    total_rows = [r for r in range(hdr + 1, ws.max_row + 1) if ws.cell(r, 1).value == "Total"]
    assert total_rows, "Totals row not found"
    tot = total_rows[0]
    # Grand total = (100k+50k) + (20k+10k) + (-5k+3k) = 178_000
    assert ws.cell(tot, last_col).value == pytest.approx(178_000.0)


def test_book2_delta_dod_empty_state():
    wb = _fresh_wb()
    _write_book2_delta_dod(wb, {}, "2026-04-18")
    assert "BOOK2 \u0394MTM - Day-over-Day" in wb.sheetnames


def test_book2_delta_dod_populated():
    data = {"book2_delta_dod": {
        "has_data": True,
        "prev_date": "2026-04-17",
        "rows": [
            {"currency": "CHF", "n_deals": 3, "mtm_today": 150_000.0,
             "mtm_prev": 100_000.0, "delta": 50_000.0},
            {"currency": "EUR", "n_deals": 2, "mtm_today": -50_000.0,
             "mtm_prev": -75_000.0, "delta": 25_000.0},
        ],
        "totals": {"mtm_today": 100_000.0, "mtm_prev": 25_000.0, "delta": 75_000.0},
    }}
    wb = _fresh_wb()
    _write_book2_delta_dod(wb, data, "2026-04-18")
    ws = wb["BOOK2 \u0394MTM - Day-over-Day"]

    # Find header row
    hdr_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Currency"]
    assert hdr_rows
    hdr = hdr_rows[0]
    assert ws.cell(hdr, 5).value == "\u0394MTM"

    # Row for CHF
    chf_rows = [r for r in range(hdr + 1, ws.max_row + 1) if ws.cell(r, 1).value == "CHF"]
    assert chf_rows
    chf = chf_rows[0]
    assert ws.cell(chf, 2).value == 3
    assert ws.cell(chf, 5).value == pytest.approx(50_000.0)

    # Totals row
    total_rows = [r for r in range(hdr + 1, ws.max_row + 1) if ws.cell(r, 1).value == "Total"]
    assert total_rows
    tot = total_rows[0]
    assert ws.cell(tot, 5).value == pytest.approx(75_000.0)


def test_book2_delta_mtd_empty_state():
    wb = _fresh_wb()
    _write_book2_delta_mtd(wb, {}, "2026-04-18")
    assert "BOOK2 \u0394MTM - MTD" in wb.sheetnames


def test_book2_delta_mtd_populated():
    data = {"book2_delta_mtd": {
        "has_data": True,
        "prev_date": "2026-03-31",
        "rows": [
            {"currency": "CHF", "n_deals": 5, "mtm_today": 200_000.0,
             "mtm_prev": 80_000.0, "delta": 120_000.0},
        ],
        "totals": {"mtm_today": 200_000.0, "mtm_prev": 80_000.0, "delta": 120_000.0},
    }}
    wb = _fresh_wb()
    _write_book2_delta_mtd(wb, data, "2026-04-18")
    ws = wb["BOOK2 \u0394MTM - MTD"]

    # Locate CHF row and verify delta
    chf_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "CHF"]
    assert chf_rows
    chf = chf_rows[0]
    assert ws.cell(chf, 5).value == pytest.approx(120_000.0)


def test_book1_realized_mtd_empty_state():
    wb = _fresh_wb()
    _write_book1_realized_mtd(wb, {}, "2026-04-18")
    assert "BOOK1 Realized - MTD" in wb.sheetnames


def test_book1_realized_mtd_populated():
    data = {"book1_realized_mtd": {
        "has_data": True,
        "month_start": "2026-04-01",
        "date_run": "2026-04-18",
        "days_counted": 13,
        "rows": [
            {"currency": "CHF", "book1_accrual": 500_000.0, "book1_ias": 50_000.0,
             "book2_mtm": -10_000.0, "book1_total": 550_000.0},
            {"currency": "EUR", "book1_accrual": 200_000.0, "book1_ias": 30_000.0,
             "book2_mtm": 5_000.0, "book1_total": 230_000.0},
        ],
        "totals": {
            "book1_accrual": 700_000.0, "book1_ias": 80_000.0,
            "book2_mtm": -5_000.0, "book1_total": 780_000.0,
        },
    }}
    wb = _fresh_wb()
    _write_book1_realized_mtd(wb, data, "2026-04-18")
    ws = wb["BOOK1 Realized - MTD"]

    # Locate CHF row and verify accrual + ias + total
    chf_rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "CHF"]
    assert chf_rows
    chf = chf_rows[0]
    assert ws.cell(chf, 2).value == pytest.approx(500_000.0)
    assert ws.cell(chf, 3).value == pytest.approx(50_000.0)
    assert ws.cell(chf, 4).value == pytest.approx(550_000.0)

    # Grand total row
    total_rows = [r for r in range(chf + 1, ws.max_row + 1) if ws.cell(r, 1).value == "Total"]
    assert total_rows
    tot = total_rows[0]
    assert ws.cell(tot, 4).value == pytest.approx(780_000.0)


def test_attribution_with_new_and_matured_deals():
    data = {"attribution": {
        "has_data": True,
        "summary": {
            "prev_nii": 1_000_000.0, "curr_nii": 1_150_000.0, "delta": 150_000.0,
            "rate_effect": 60_000.0, "spread_effect": 20_000.0,
            "time_effect": 10_000.0,
            "new_deal_effect": 80_000.0, "matured_deal_effect": -20_000.0,
            "n_existing": 42, "n_new": 3, "n_matured": 2,
        },
        "by_currency": {"CHF": {"prev_nii": 600_000, "curr_nii": 700_000, "rate_effect": 35_000}},
        "new_deals": [
            {"deal_id": "N-001", "counterparty": "UBS", "currency": "CHF",
             "product": "IAM/LD", "pnl": 50_000.0, "nominal": 25_000_000.0},
            {"deal_id": "N-002", "counterparty": "CS", "currency": "EUR",
             "product": "BND", "pnl": 30_000.0, "nominal": 15_000_000.0},
        ],
        "matured_deals": [
            {"deal_id": "M-001", "counterparty": "Zurich", "currency": "CHF",
             "product": "IAM/LD", "pnl_lost": 15_000.0, "nominal": 10_000_000.0},
            {"deal_id": "M-002", "counterparty": "Swiss Re", "currency": "USD",
             "product": "BND", "pnl_lost": 5_000.0, "nominal": 5_000_000.0},
        ],
    }}
    wb = _fresh_wb()
    _write_attribution(wb, data, "2026-04-18")
    ws = wb["Attribution Waterfall"]

    # New-deals section — locate header row with "Deal ID"
    deal_id_header_rows = [
        r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Deal ID"
    ]
    # Two sections (new + matured) both open with a "Deal ID" header
    assert len(deal_id_header_rows) == 2

    # First header is new deals — verify Deal ID values on the two rows after it
    new_hdr = deal_id_header_rows[0]
    assert ws.cell(new_hdr + 1, 1).value == "N-001"
    assert ws.cell(new_hdr + 2, 1).value == "N-002"
    assert ws.cell(new_hdr + 1, 6).value == pytest.approx(50_000.0)

    # Second header is matured
    mat_hdr = deal_id_header_rows[1]
    assert ws.cell(mat_hdr + 1, 1).value == "M-001"
    assert ws.cell(mat_hdr + 2, 1).value == "M-002"
    assert ws.cell(mat_hdr + 1, 6).value == pytest.approx(15_000.0)


def test_attribution_empty_state():
    wb = _fresh_wb()
    _write_attribution(wb, {}, "2026-04-18")
    assert "Attribution Waterfall" in wb.sheetnames


def test_attribution_mom_empty_state():
    wb = _fresh_wb()
    _write_attribution_mom(wb, {}, "2026-04-18")
    assert "P&L Explain (MoM)" in wb.sheetnames


def test_attribution_mom_populated():
    data = {"attribution_mom": {
        "has_data": True,
        "summary": {
            "prev_nii": 10_000_000.0, "curr_nii": 13_080_000.0, "delta": 3_080_000.0,
            "rate_effect": 900_000.0, "spread_effect": -200_000.0,
            "time_effect": 50_000.0,
            "new_deal_effect": 1_800_000.0, "matured_deal_effect": -500_000.0,
            "n_existing": 260, "n_new": 27, "n_matured": 14,
            "prev_date": "2026-03-31", "curr_date": "2026-04-18",
        },
        "by_currency": {"CHF": {"prev_nii": 6_000_000, "curr_nii": 7_540_000, "rate_effect": 500_000}},
        "new_deals": [
            {"deal_id": "304200", "counterparty": "UBS", "currency": "CHF",
             "product": "IAM/LD", "pnl": 350_000.0, "nominal": 80_000_000.0},
        ],
        "matured_deals": [
            {"deal_id": "200100", "counterparty": "Credit Suisse", "currency": "CHF",
             "product": "BND", "pnl_lost": 180_000.0, "nominal": 40_000_000.0},
        ],
    }}
    wb = _fresh_wb()
    _write_attribution_mom(wb, data, "2026-04-18")
    ws = wb["P&L Explain (MoM)"]

    # Title should mention month-start anchor
    title = ws.cell(1, 1).value or ""
    assert "Month-over-Month" in title
    assert "2026-03-31" in title

    # New-deals section present
    deal_id_header_rows = [
        r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Deal ID"
    ]
    assert len(deal_id_header_rows) == 2  # new + matured
