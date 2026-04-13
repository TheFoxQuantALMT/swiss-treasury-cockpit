"""Generate data_dictionary.xlsx for all input Excel files."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap = Alignment(wrap_text=True, vertical="top")


def make_sheet(ws, title, headers, rows, col_widths=None):
    ws.title = title
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap
            cell.border = thin_border
            if len(row) > 3 and row[3] == "Yes" and c == 1:
                cell.font = Font(bold=True)
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"


COL_HEADERS = [
    "Column Name", "Internal Name", "Data Type",
    "Required", "Valid Values", "Default", "Description",
]
COL_WIDTHS = [22, 18, 10, 9, 30, 12, 50]

# ── INDEX ────────────────────────────────────────────────────────────
ws_idx = wb.active
make_sheet(ws_idx, "INDEX",
    ["File Name", "Sheet Name", "Parser Function", "Type", "Primary Key(s)", "Description"],
    [
        ("deals.xlsx", "Deals", "parse_deals()", "Core", "deal_id",
         "Deal master data: product, amount, rates, maturity, counterparty, FTP"),
        ("rate_schedule.xlsx", "Schedule", "parse_schedule()", "Core", "deal_id + YYYY/MM",
         "Monthly nominal balance schedules per deal (wide format)"),
        ("wirp.xlsx", "WIRP", "parse_wirp_ideal()", "Core", "index + meeting_date",
         "Market-implied rate expectations by currency (WIRP Bloomberg)"),
        ("reference_table.xlsx", "Reference", "parse_reference_table()", "Core", "counterparty",
         "Counterparty credit ratings, HQLA level, country codes"),
        ("budget.xlsx", "Budget", "parse_budget()", "Optional", "currency + month",
         "Monthly NII budget targets per currency and perimeter"),
        ("scenarios.xlsx", "Scenarios", "parse_scenarios()", "Optional", "scenario + tenor",
         "BCBS 368 tenor-dependent rate shock definitions (basis points)"),
        ("nmd_profiles.xlsx", "NMD", "parse_nmd_profiles()", "Optional",
         "product + currency + direction",
         "Non-Maturing Deposit behavioral decay parameters"),
        ("limits.xlsx", "Limits", "parse_limits()", "Optional", "metric + currency",
         "Board-approved NII/EVE/concentration limits"),
        ("alert_thresholds.xlsx", "Thresholds", "parse_alert_thresholds()", "Optional", "currency",
         "Per-currency alert threshold overrides"),
        ("liquidity_rate_schedule.xlsx", "Liquidity", "parse_liquidity_schedule()", "Optional",
         "deal_id + date cols",
         "Daily (90d) + monthly cash flow projections per deal"),
        ("custom_scenarios.xlsx", "(first sheet)", "parse_custom_scenarios()", "Optional",
         "scenario + tenor",
         "User-defined stress test scenarios (shocks in bps)"),
    ],
    [22, 14, 24, 9, 28, 60],
)

# ── deals.xlsx ───────────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "deals", COL_HEADERS, [
    ("deal_id", "Dealid", "int", "Yes", "Numeric, non-null", "-",
     "Unique deal identifier. Join key with rate_schedule.xlsx"),
    ("product", "Product", "str", "Yes", "IAM/LD, BND, FXS, IRS, IRS-MTM, HCD", "-",
     "Product type. Determines which rate column is used as RateRef"),
    ("currency", "Currency", "str", "Yes", "CHF, EUR, USD, GBP", "-",
     "Deal currency. Filtered against SUPPORTED_CURRENCIES"),
    ("direction", "Direction", "str", "Yes", "B, L, D, S", "-",
     "B=Bond, L=Lend/Loan, D=Deposit, S=Sell"),
    ("book", "IAS Book", "str", "Yes", "BOOK1, BOOK2", "-",
     "IAS book designation. BOOK2 = MTM swap inventory"),
    ("amount", "Amount", "float", "No", "Any signed float", "0.0",
     "Outstanding balance amount"),
    ("client_rate", "Clientrate", "float", "No", "Decimal (-0.50 to +0.50)", "0.0",
     "Contractual client rate in decimal (e.g., 0.015 = 1.5%). Warns if >50%"),
    ("eq_ois_rate", "EqOisRate", "float", "No", "Decimal", "0.0",
     "BD-1 equivalent OIS rate in decimal. Used as RateRef for IAM/LD, FXS"),
    ("ytm", "YTM", "float", "No", "Decimal", "0.0",
     "Yield to maturity net of credit spread. Used as RateRef for BND"),
    ("coc_rate", "CocRate", "float", "No", "Decimal", "0.0",
     "Deal-specific cost of carry rate"),
    ("spread", "Spread", "float", "No", "Decimal (not bps)", "0.0",
     "Credit spread in decimal form"),
    ("floating_index", "Floating Rates Short Name", "str", "No",
     'SARON, ESTR, SOFR, SONIA, (empty)', '""',
     "RFR index name. Empty = fixed-rate deal. Non-empty -> is_floating=True"),
    ("trade_date", "Tradedate", "date", "No", "ISO 8601, dayfirst=True", "null",
     "Trade date"),
    ("value_date", "Valuedate", "date", "No", "ISO 8601, dayfirst=True", "null",
     "Value / settlement date"),
    ("maturity_date", "Maturitydate", "date", "Yes", "ISO 8601, dayfirst=True", "-",
     "Maturity date. Rows with null maturity are dropped"),
    ("strategy_ias", "Strategy IAS", "str", "No", "Any string", "null",
     "Hedge designation label. Groups deals into hedge relationships. "
     "Deals with same value form a pair (hedged items vs instruments by product)"),
    ("hedge_type", "hedge_type", "str", "No", "cash_flow, fair_value", "cash_flow",
     "Hedge relationship type per IAS39/IFRS9. Set on all deals sharing a strategy_ias"),
    ("ias_standard", "ias_standard", "str", "No", "IAS39, IFRS9", "IFRS9",
     "Accounting standard for effectiveness test. IAS39=dollar-offset, IFRS9=R-squared"),
    ("designation_date", "designation_date", "date", "No", "ISO 8601", "null",
     "Date the hedge relationship was formally designated"),
    ("perimeter", "Perimetre TOTAL", "str", "Yes", "CC, WM, CIB", "CC",
     "Business perimeter. Defaults to CC if invalid"),
    ("counterparty", "Counterparty", "str", "No", "Any string", '""',
     "Counterparty code / name"),
    ("ftp", "FTP", "float", "No", "Decimal", "null",
     "Funds Transfer Pricing rate. Enables 3-way margin split"),
    ("pay_receive", "pay_receive", "str", "No", "PAY, RECEIVE", "null",
     "BOOK2 IRS only: pay or receive fixed leg"),
    ("notional", "notional", "float", "No", "Any float", "0.0",
     "BOOK2 IRS only: swap notional amount"),
    ("last_fixing_date", "last_fixing_date", "date", "No", "Date", "null",
     "BOOK2 IRS only: last RFR fixing date"),
    ("next_fixing_date", "next_fixing_date", "date", "No", "Date", "null",
     "BOOK2 IRS only: next RFR fixing date"),
], COL_WIDTHS)

# ── rate_schedule.xlsx ────────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "schedule", COL_HEADERS, [
    ("deal_id", "Dealid", "int", "Yes", "Numeric, non-null", "-",
     "Deal identifier. Must match deal_id in deals.xlsx"),
    ("direction", "Direction", "str", "Yes", "B, L, D, S", "-",
     "Must match direction in deals.xlsx"),
    ("currency", "Currency", "str", "Yes", "CHF, EUR, USD, GBP", "-",
     "Must match currency in deals.xlsx"),
    ("rate_type", "Rate Type", "str", "No", "F (Fixed), V (Variable)", "-",
     "F=Fixed rate, V=Variable/floating rate"),
    ("YYYY/MM", "(month columns)", "float", "No",
     "Any float (negative = liability)", "0",
     "Monthly outstanding nominal balance. Columns span the deal horizon "
     "(e.g., 2026/04 through 2031/03). Zero = matured or not yet started. "
     "Positive for assets, negative for liabilities"),
], COL_WIDTHS)

# ── wirp.xlsx ────────────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "wirp", COL_HEADERS, [
    ("index", "Indice", "str", "Yes", "CHFSON, EUREST, USSOFR, GBPOIS", "-",
     "OIS index code. Maps to currency via CURRENCY_TO_OIS"),
    ("meeting_date", "Meeting", "date", "Yes", "Non-null date, dayfirst=True", "-",
     "Central bank meeting date (SNB, ECB, Fed, BoE)"),
    ("rate", "Rate", "float", "No", "Decimal, warns if |v|>0.20", "-",
     "Market-implied policy rate at that meeting in decimal"),
    ("change_bps", "Hike / Cut", "float", "No", "Basis points", "-",
     "Expected rate change in basis points at that meeting"),
], COL_WIDTHS)

# ── reference_table.xlsx ─────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "reference_table", COL_HEADERS, [
    ("counterparty", "counterparty", "str", "Yes", "Any string", "-",
     "Counterparty code. Must match counterparty field in deals.xlsx"),
    ("rating", "rating", "str", "No", "Standard credit ratings (AAA..D)", "NR",
     "Credit rating. NR = Not Rated"),
    ("hqla_level", "hqla_level", "str", "No", "L1, L2A, L2B, Non-HQLA", "Non-HQLA",
     "HQLA classification for LCR/SNB reserves"),
    ("country", "country", "str", "No", "ISO 3166-1 alpha-2", "XX",
     "Country code of the counterparty"),
], COL_WIDTHS)

# ── budget.xlsx ──────────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "budget", COL_HEADERS, [
    ("currency", "currency", "str", "Yes", "CHF, EUR, USD, GBP", "-",
     "Currency for the budget line"),
    ("month", "month", "str", "Yes", "YYYY-MM format", "-",
     "Budget month (e.g., 2026-04)"),
    ("budget_nii", "budget_nii", "float", "Yes", "Any float", "-",
     "Budgeted Net Interest Income for that currency/month"),
    ("budget_nominal", "budget_nominal", "float", "No", "Any float", "0.0",
     "Budgeted outstanding nominal amount"),
    ("budget_rate", "budget_rate", "float", "No", "Decimal", "0.0",
     "Budgeted average rate target"),
    ("perimeter", "perimeter", "str", "No", "CC, WM, CIB", "CC",
     "Business perimeter for the budget line"),
    ("product", "product", "str", "No", "Any string", "null",
     "Product filter (optional)"),
], COL_WIDTHS)

# ── scenarios.xlsx ───────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "scenarios", COL_HEADERS, [
    ("scenario", "scenario", "str", "Yes",
     "parallel_up, parallel_down, short_up, short_down, steepener, flattener", "-",
     "BCBS 368 scenario name. 6 standard non-parallel rate shocks"),
    ("tenor", "tenor", "str", "Yes",
     "O/N, 3M, 6M, 1Y, 2Y, 3Y, 5Y, 10Y, 20Y, 30Y", "-",
     "Tenor point on the yield curve (BCBS 368 convention)"),
    ("CHF", "CHF", "float", "No", "Basis points (e.g., 200 = +200bp)", "0.0",
     "Rate shock in bps for CHF at this tenor"),
    ("EUR", "EUR", "float", "No", "Basis points", "0.0",
     "Rate shock in bps for EUR at this tenor"),
    ("USD", "USD", "float", "No", "Basis points", "0.0",
     "Rate shock in bps for USD at this tenor"),
    ("GBP", "GBP", "float", "No", "Basis points", "0.0",
     "Rate shock in bps for GBP at this tenor"),
], COL_WIDTHS)

# ── nmd_profiles.xlsx ────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "nmd_profiles", COL_HEADERS, [
    ("product", "product", "str", "Yes", "e.g., IAM/LD", "-",
     "Product type for NMD classification"),
    ("currency", "currency", "str", "Yes", "CHF, EUR, USD, GBP", "-",
     "Currency of the NMD portfolio"),
    ("direction", "direction", "str", "Yes", "B, L, D, S", "-",
     "Direction (typically L for sight deposits)"),
    ("tier", "tier", "str", "No", "core, volatile, term", "core",
     "Deposit stability tier. Core = most stable"),
    ("behavioral_maturity_years", "behavioral_maturity_years", "float", "No",
     "Positive float (years)", "5.0",
     "Modeled repricing horizon in years"),
    ("decay_rate", "decay_rate", "float", "No", "0 to 1", "0.15",
     "Annual exponential runoff rate. Applied as exp(-decay * t)"),
    ("deposit_beta", "deposit_beta", "float", "No", "0 to 1", "0.5",
     "Rate passthrough coefficient (delta_client / delta_OIS). 0=no passthrough, 1=full"),
    ("floor_rate", "floor_rate", "float", "No", "Decimal", "0.0",
     "Minimum client rate floor (e.g., 0.0 = zero floor)"),
], COL_WIDTHS)

# ── limits.xlsx ──────────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "limits", COL_HEADERS, [
    ("metric", "metric", "str", "Yes",
     "nii_sensitivity_50bp, nii_at_risk_worst, eve_change_200bp, "
     "eve_change_worst, concentration_hhi", "-",
     "Risk metric to limit"),
    ("currency", "currency", "str", "No", "CHF, EUR, USD, GBP, ALL", "ALL",
     "Currency scope. ALL = applies to total"),
    ("limit_value", "limit_value", "float", "Yes", "Any float", "-",
     "Limit threshold value (absolute or relative per limit_type)"),
    ("warning_pct", "warning_pct", "float", "No", "0-100 (percentage)", "80.0",
     "Yellow/warning alert at this % of limit utilization"),
    ("limit_type", "limit_type", "str", "No", "absolute, relative", "absolute",
     "How to interpret limit_value"),
], COL_WIDTHS)

# ── alert_thresholds.xlsx ────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "alert_thresholds", COL_HEADERS, [
    ("currency", "currency", "str", "Yes", "CHF, EUR, USD, GBP, ALL", "-",
     "Currency scope. ALL sets global defaults; specific currency overrides"),
    ("annual_nii_floor", "annual_nii_floor", "float", "No", "Any float", "null",
     "Minimum acceptable annual NII. Triggers alert if NII falls below"),
    ("mom_delta_pct", "mom_delta_pct", "float", "No", "Percentage (e.g., 5.0)", "null",
     "Month-on-month NII change % that triggers an alert"),
    ("ccy_concentration_pct", "ccy_concentration_pct", "float", "No",
     "Percentage (e.g., 40.0)", "null",
     "Currency concentration % above which an alert fires"),
    ("shock_sensitivity_limit", "shock_sensitivity_limit", "float", "No",
     "Any float", "null",
     "+50bp NII sensitivity limit. Alert if delta exceeds this"),
], COL_WIDTHS)

# ── liquidity_rate_schedule.xlsx ──────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "liquidity_schedule", COL_HEADERS, [
    ("deal_id", "Dealid", "int", "Yes", "Numeric, non-null", "-",
     "Deal identifier. Should match deal_id in deals.xlsx"),
    ("direction", "Direction", "str", "Yes", "B, L, D, S", "-",
     "Deal direction"),
    ("currency", "Currency", "str", "Yes", "CHF, EUR, USD, GBP", "-",
     "Deal currency"),
    ("YYYY/MM/DD", "(daily columns)", "float", "No", "Cash flow amount", "0.0",
     "Daily cash flows for the first ~90 days (e.g., 2026/04/07). "
     "Interest + principal. Positive = inflow, negative = outflow"),
    ("YYYY/MM", "(monthly columns)", "float", "No", "Cash flow amount", "0.0",
     "Monthly aggregated cash flows beyond 90 days (e.g., 2026/07). "
     "Same sign convention as daily"),
], COL_WIDTHS)

# ── custom_scenarios.xlsx ────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "custom_scenarios", COL_HEADERS, [
    ("scenario", "scenario", "str", "Yes",
     "Any user-defined name (e.g., SNB_reversal)", "-",
     "Custom scenario name. Converted to BCBS-compatible format for engine"),
    ("tenor", "tenor", "float", "Yes",
     "Years: 0.25, 0.5, 1, 2, 3, 5, 7, 10, 15, 20", "-",
     "Tenor point in years (BCBS 368 convention)"),
    ("CHF", "CHF", "float", "No", "Basis points (e.g., -50 = -50bp cut)", "0.0",
     "Rate shock for CHF at this tenor in basis points"),
    ("EUR", "EUR", "float", "No", "Basis points", "0.0",
     "Rate shock for EUR at this tenor in basis points"),
    ("USD", "USD", "float", "No", "Basis points", "0.0",
     "Rate shock for USD at this tenor in basis points"),
    ("GBP", "GBP", "float", "No", "Basis points", "0.0",
     "Rate shock for GBP at this tenor in basis points"),
], COL_WIDTHS)

# ── Constants & Enums ────────────────────────────────────────────────
ws = wb.create_sheet()
make_sheet(ws, "Constants & Enums",
    ["Category", "Key", "Value", "Description"],
    [
        ("Currencies", "SUPPORTED_CURRENCIES", "CHF, EUR, USD, GBP",
         "Only these 4 currencies are processed"),
        ("", "", "", ""),
        ("OIS Index Mapping", "CHF", "CHFSON", "Swiss Average Rate Overnight"),
        ("OIS Index Mapping", "EUR", "EUREST", "Euro Short-Term Rate"),
        ("OIS Index Mapping", "USD", "USSOFR", "Secured Overnight Financing Rate"),
        ("OIS Index Mapping", "GBP", "GBPOIS", "Sterling Overnight Index Average"),
        ("", "", "", ""),
        ("Floating Index to WASP", "SARON", "CHFSON", "Swiss SARON"),
        ("Floating Index to WASP", "ESTR", "EUREST", "Euro EUR short-term rate"),
        ("Floating Index to WASP", "SOFR", "USSOFR", "US SOFR"),
        ("Floating Index to WASP", "SONIA", "GBPOIS", "UK SONIA"),
        ("", "", "", ""),
        ("Product to RateRef", "IAM/LD", "EqOisRate",
         "Deposits/loans use BD-1 equivalent OIS rate"),
        ("Product to RateRef", "BND", "YTM", "Bonds use yield to maturity"),
        ("Product to RateRef", "FXS", "EqOisRate",
         "FX swaps use BD-1 equivalent OIS rate"),
        ("Product to RateRef", "IRS", "Clientrate",
         "Interest rate swaps use contractual rate"),
        ("Product to RateRef", "IRS-MTM", "Clientrate",
         "Mark-to-market IRS use contractual rate"),
        ("Product to RateRef", "HCD", "Clientrate",
         "Hedge-designated deals use contractual rate"),
        ("", "", "", ""),
        ("Day Count (MM)", "CHF", "360", "Act/360"),
        ("Day Count (MM)", "EUR", "360", "Act/360"),
        ("Day Count (MM)", "USD", "360", "Act/360"),
        ("Day Count (MM)", "GBP", "365", "Act/365"),
        ("", "", "", ""),
        ("RFR Lookback", "CHF (SARON)", "2 business days",
         "ISDA 2021 observation shift"),
        ("RFR Lookback", "GBP (SONIA)", "5 business days",
         "ISDA 2021 observation shift"),
        ("", "", "", ""),
        ("Directions", "B", "Bond", "Bond purchase / asset holding"),
        ("Directions", "L", "Lend / Loan",
         "Lending (liability side, negative nominal)"),
        ("Directions", "D", "Deposit", "Deposit placement (asset side)"),
        ("Directions", "S", "Sell", "Bond sale / short position"),
        ("", "", "", ""),
        ("Books", "BOOK1", "Accrual book", "Standard P&L accrual (NII)"),
        ("Books", "BOOK2", "MTM book",
         "Mark-to-market (IRS swaps, fair value)"),
        ("", "", "", ""),
        ("Perimeters", "CC", "Corporate Center",
         "Default perimeter for ALM treasury"),
        ("Perimeters", "WM", "Wealth Management",
         "Private banking / WM business unit"),
        ("Perimeters", "CIB", "Corporate & Investment Banking",
         "CIB business unit"),
        ("", "", "", ""),
        ("NMD Tiers", "core", "Core deposits",
         "Stable, long-behavioral-maturity deposits"),
        ("NMD Tiers", "volatile", "Volatile deposits",
         "Rate-sensitive, shorter behavioral maturity"),
        ("NMD Tiers", "term", "Term deposits",
         "Contractual maturity deposits"),
        ("", "", "", ""),
        ("BCBS 368 Scenarios", "parallel_up", "+200bp parallel shift",
         "All tenors shift up equally"),
        ("BCBS 368 Scenarios", "parallel_down", "-200bp parallel shift",
         "All tenors shift down equally"),
        ("BCBS 368 Scenarios", "short_up", "+300bp short-end rise",
         "Short tenors rise, long tenors less affected"),
        ("BCBS 368 Scenarios", "short_down", "-300bp short-end drop",
         "Short tenors drop, long tenors less affected"),
        ("BCBS 368 Scenarios", "steepener", "Curve steepening",
         "Short rates down, long rates up"),
        ("BCBS 368 Scenarios", "flattener", "Curve flattening",
         "Short rates up, long rates down"),
        ("", "", "", ""),
        ("Rate Convention", "All rates", "Decimal",
         "Rates stored as decimal (0.015 = 1.5%), never percent"),
        ("Rate Convention", "Shocks", "Basis points",
         "Scenario shocks in bps (200 = +200bp = +2%)"),
    ],
    [22, 20, 28, 50],
)

out = "tests/fixtures/ideal_input/data_dictionary.xlsx"
wb.save(out)
print(f"Created {out}")
print(f"Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
